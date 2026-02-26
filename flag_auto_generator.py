import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import time
import zipfile

import ttkbootstrap as tb
from ttkbootstrap import ttk
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


AUTO_DATA_START_ROW_DEFAULT = 230
AUTO_DATA_MAX_ITEMS = 100


def pick_file(title: str, filetypes, parent=None):
    if parent is None:
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        root.destroy()
        return path
    return filedialog.askopenfilename(parent=parent, title=title, filetypes=filetypes)


def pick_save_path(title: str, defaultextension: str, filetypes, parent=None):
    if parent is None:
        root = tk.Tk()
        root.withdraw()
        path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=defaultextension,
            filetypes=filetypes,
        )
        root.destroy()
        return path
    return filedialog.asksaveasfilename(
        parent=parent,
        title=title,
        defaultextension=defaultextension,
        filetypes=filetypes,
    )


def _save_workbook_atomic(wb, out_path: str, parent=None) -> str:
    out_path = os.path.abspath(out_path)
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    base, ext = os.path.splitext(out_path)
    ext = ext or ".xlsx"
    tmp_path = f"{base}.tmp{ext}"

    while True:
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, out_path)
            return out_path
        except PermissionError as e:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
            if parent is None:
                raise
            retry = messagebox.askretrycancel(
                "保存に失敗",
                "出力先ファイルに書き込めません。\n"
                "Excelで出力先ファイルを開いている場合は閉じてから「再試行」を押してください。\n\n"
                f"出力先:\n{out_path}\n\n"
                f"詳細:\n{e}",
                parent=parent,
            )
            if not retry:
                raise
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass


def _mark_workbook_for_full_recalc(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _safe_call(func, *args, default=None, **kwargs):
    try:
        return func(*args, **kwargs)
    except Exception:
        return default


def _close_workbook_quietly(wb):
    if wb is None:
        return
    _safe_call(wb.close)


def _restore_package_parts_from_source(source_xlsx_path: str, target_xlsx_path: str):
    prefixes = (
        "xl/drawings/",
        "xl/externalLinks/",
    )

    source_path = os.path.abspath(source_xlsx_path)
    target_path = os.path.abspath(target_xlsx_path)
    temp_path = f"{target_path}.pkgfix"

    try:
        with zipfile.ZipFile(source_path, "r") as src_zip, zipfile.ZipFile(
            target_path, "r"
        ) as dst_zip:
            src_names = set(src_zip.namelist())
            dst_names = set(dst_zip.namelist())
            restore_names = {
                name
                for name in src_names
                if any(name.startswith(prefix) for prefix in prefixes)
            }
            if not restore_names:
                return False

            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
                for name in dst_zip.namelist():
                    if any(name.startswith(prefix) for prefix in prefixes):
                        if name in restore_names:
                            out_zip.writestr(name, src_zip.read(name))
                        continue
                    out_zip.writestr(name, dst_zip.read(name))

                for name in sorted(restore_names):
                    if name in dst_names:
                        continue
                    out_zip.writestr(name, src_zip.read(name))

        os.replace(temp_path, target_path)
        return True
    except Exception as e:
        _safe_call(os.remove, temp_path)
        print(f"[info] パッケージ復元をスキップしました: {e}")
        return False


def _force_excel_recalc_and_save(xlsx_path: str):
    try:
        import win32com.client  # type: ignore
    except Exception as e:
        print(f"[warn] Excel強制再計算をスキップしました（pywin32未導入）: {e}")
        return False

    abs_path = os.path.abspath(xlsx_path)
    last_error = None

    for attempt in range(3):
        excel_app = None
        excel_wb = None
        try:
            excel_app = win32com.client.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            _safe_call(setattr, excel_app, "AskToUpdateLinks", False)

            excel_wb = excel_app.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=False)

            _safe_call(setattr, excel_wb, "ForceFullCalculation", True)

            worksheets = _safe_call(lambda: list(excel_wb.Worksheets), default=[])
            for sheet in worksheets:
                _safe_call(setattr, sheet, "EnableCalculation", True)
                _safe_call(sheet.UsedRange.Calculate)

            _safe_call(excel_wb.RefreshAll)
            _safe_call(excel_app.CalculateFull)
            excel_app.CalculateFullRebuild()
            save_as_kwargs = {
                "Filename": abs_path,
                "FileFormat": 51,
                "ConflictResolution": 2,
                "Local": True,
            }
            if _safe_call(excel_wb.SaveAs, **save_as_kwargs) is None:
                excel_wb.Save()
            return True
        except Exception as e:
            last_error = e
            time.sleep(0.4 * (attempt + 1))
        finally:
            if excel_wb is not None:
                _safe_call(excel_wb.Close, SaveChanges=True)
            if excel_app is not None:
                _safe_call(excel_app.Quit)

    print(f"[info] Excel強制再計算をスキップしました（出力ファイルは作成済み）: {last_error}")
    return False


def _try_extract_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        try:
            return int(value)
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(s)
        except Exception:
            m = re.search(r"\d+", s)
            if not m:
                return None
            try:
                return int(m.group(0))
            except Exception:
                return None
    return None


def _resolve_measure_no(value, row: int, measure_row_min: int, measure_row_step: int):
    if isinstance(value, str) and value.startswith("="):
        if measure_row_step <= 0:
            return None

        row_offset = row - measure_row_min
        if row_offset < 0:
            return None
        if row_offset % measure_row_step != 0:
            return None

        return (row_offset // measure_row_step) + 1

    return _try_extract_int(value)


def _normalize_measure_no_key(value):
    n = _try_extract_int(value)
    if n is not None:
        return n
    if value is None:
        return ""
    return str(value).strip()


def _normalize_measure_to_index_map(raw_map: dict):
    normalized = {}
    if not isinstance(raw_map, dict):
        return normalized
    for measure_no, data_index in raw_map.items():
        key = _normalize_measure_no_key(measure_no)
        if key == "":
            continue
        index_value = _try_extract_int(data_index)
        if index_value is None or index_value < 1 or index_value > AUTO_DATA_MAX_ITEMS:
            continue
        normalized[key] = index_value
    return normalized


def _build_auto_data_formula(col_letter: str, data_start_row: int, data_index: int, sep: str):
    source_row = data_start_row + data_index - 1
    return (
        f'IFERROR(IF({col_letter}{source_row}=""{sep}""{sep}{col_letter}{source_row}){sep}"")'
    )


def _is_empty_cell_value(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _can_overwrite_with_formula(value):
    if _is_empty_cell_value(value):
        return True
    if isinstance(value, str) and value.startswith("="):
        return True
    return False


def _normalize_single_cell_array_formulas_in_column(
    ws,
    col_letter: str,
    *,
    row_start: int = 1,
    row_end: int | None = None,
):
    col_idx = column_index_from_string(col_letter)
    max_row = ws.max_row or row_start
    end_row = max_row if row_end is None else min(row_end, max_row)
    rewritten = 0

    start_row = max(row_start, 1)
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row, col_idx)
        value = cell.value
        if not isinstance(value, ArrayFormula):
            continue

        formula_ref = str(value.ref or "").replace("$", "")
        cell_ref = cell.coordinate.replace("$", "")
        if formula_ref != cell_ref:
            continue

        formula_text = str(value.text or "").strip()
        if not formula_text:
            continue

        cell.value = f"={formula_text.lstrip('=')}"
        rewritten += 1

    return rewritten


def build_request_formulas(xlsx_path: str, out_path: str, cfg: dict, *, parent=None):
    sheet_name = cfg["sheet_name"]
    measure_no_col = cfg.get("measure_no_col", "A")
    measure_row_min = int(cfg.get("measure_row_min", 11))
    measure_row_step = int(cfg.get("measure_row_step", 3))
    tool_start_row = int(cfg.get("tool_start_row", 200))
    measure_row_max = int(cfg.get("measure_row_max", tool_start_row - 4))
    summary_row_min = int(cfg.get("summary_row_min", measure_row_min))
    summary_row_max = int(cfg.get("summary_row_max", measure_row_max))
    summary_row_step = int(cfg.get("summary_row_step", measure_row_step))
    formula_arg_sep = str(cfg.get("formula_arg_sep", ",")).strip() or ","
    flag_col_start = column_index_from_string("L")
    flag_col_end = column_index_from_string("SR")

    tool_name_col = cfg.get("tool_name_col", "E")
    tool_row_step = int(cfg.get("tool_row_step", measure_row_step))
    auto_data_start_row = int(cfg.get("auto_data_start_row", AUTO_DATA_START_ROW_DEFAULT))
    measure_no_to_data_index = _normalize_measure_to_index_map(
        cfg.get("measure_no_to_data_index", {})
    )

    tools = cfg["tools"]
    tool_to_measure_nos = cfg["tool_to_measure_nos"]

    # SEQUENCE関数の第1引数として使用する値を計算（工具開始行-4）
    sequence_count = tool_start_row - 4

    wb = load_workbook(xlsx_path)
    wb_values = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"シート '{sheet_name}' が見つかりません。存在: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    ws_values = wb_values[sheet_name]

    # 1) 測定No(A列の整数) → 行番号
    no_col = column_index_from_string(measure_no_col)
    measure_no_to_row = {}
    measure_row_to_no = {}
    max_r = min(measure_row_max, ws.max_row or measure_row_max)
    for r in range(measure_row_min, max_r + 1):
        v = ws_values.cell(r, no_col).value
        if v is None:
            v = ws.cell(r, no_col).value
        if v is None:
            continue
        no = _resolve_measure_no(v, r, measure_row_min, measure_row_step)
        if no is None:
            continue
        measure_no_to_row[no] = r
        measure_row_to_no[r] = no

    # 2) 工具名を書き込み & 工具名 → 行番号
    tool_row = {}
    tool_name_c = column_index_from_string(tool_name_col)
    r = tool_start_row
    for tool in tools:
        tool_cell = ws.cell(r, tool_name_c)
        tool_cell.value = tool
        tool_row[tool] = r
        r += tool_row_step

    # 2.5) 自動測定データ開始行のE列に案内を記入
    auto_data_label_cell = ws.cell(auto_data_start_row, tool_name_c)
    auto_data_label_cell.value = '測定結果貼付は230行から'

    # 3) 測定行ごとに参照すべき工具行（逆引き）
    measure_row_to_tool_rows = {}
    missing_nos = []
    for tool, nos in tool_to_measure_nos.items():
        if tool not in tool_row:
            continue
        tr = tool_row[tool]
        for no in nos:
            n = _try_extract_int(no)
            if n is None:
                continue
            mr = measure_no_to_row.get(n)
            if mr is None:
                missing_nos.append((tool, n))
                continue
            measure_row_to_tool_rows.setdefault(mr, []).append(tr)

    # 4) 1～3行目に新しい数式を投入、測定行に依頼セルを設定
    written = 0
    target_found = 0
    for col_idx in range(flag_col_start, flag_col_end + 1):
        col_letter = get_column_letter(col_idx)
        
        # 1行目: SEQUENCE(sequence_count,1,11,3) - measure_row_minから開始
        formula_row1 = (
            f"=SUMPRODUCT((IFERROR(FILTER(INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min}{formula_arg_sep}{measure_row_step})){formula_arg_sep}"
            f"INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min}{formula_arg_sep}{measure_row_step}))<>\"\"){formula_arg_sep}\"\")<>\"\")*1)"
        )
        row1_cell = ws.cell(1, col_idx)
        if _is_empty_cell_value(row1_cell.value):
            row1_cell.value = formula_row1

        # 2行目: SEQUENCE(sequence_count,1,12,3) - measure_row_min+1から開始
        formula_row2 = (
            f"=SUMPRODUCT((IFERROR(FILTER(INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min + 1}{formula_arg_sep}{measure_row_step})){formula_arg_sep}"
            f"INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min + 1}{formula_arg_sep}{measure_row_step}))<>\"\"){formula_arg_sep}\"\")<>\"\")*1)"
        )
        row2_cell = ws.cell(2, col_idx)
        if _is_empty_cell_value(row2_cell.value):
            row2_cell.value = formula_row2

        # 3行目: SEQUENCE(sequence_count,1,13,3) - measure_row_min+2から開始
        formula_row3 = (
            f"=SUMPRODUCT((IFERROR(FILTER(INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min + 2}{formula_arg_sep}{measure_row_step})){formula_arg_sep}"
            f"INDEX({col_letter}:{col_letter}{formula_arg_sep}"
            f"SEQUENCE({sequence_count}{formula_arg_sep}1{formula_arg_sep}{measure_row_min + 2}{formula_arg_sep}{measure_row_step}))<>\"\"){formula_arg_sep}\"\")<>\"\")*1)"
        )
        row3_cell = ws.cell(3, col_idx)
        if _is_empty_cell_value(row3_cell.value):
            row3_cell.value = formula_row3

        # デバッグ用ログ出力
        print(f"列 {col_letter} 1行目: {formula_row1}")
        print(f"列 {col_letter} 2行目: {formula_row2}")
        print(f"列 {col_letter} 3行目: {formula_row3}")

        # 測定行に依頼セルを設定
        for mr, tool_rows in measure_row_to_tool_rows.items():
            conds = formula_arg_sep.join([f'{col_letter}${tr}<>""' for tr in tool_rows])
            measure_no = measure_row_to_no.get(mr)
            data_index = measure_no_to_data_index.get(measure_no)
            target_found += 1
            target_cell = ws.cell(mr, col_idx)
            if not _is_empty_cell_value(target_cell.value):
                continue
            if data_index is not None:
                auto_formula = _build_auto_data_formula(
                    col_letter=col_letter,
                    data_start_row=auto_data_start_row,
                    data_index=data_index,
                    sep=formula_arg_sep,
                )
                target_cell.value = f'=IF(OR({conds}),"依頼",{auto_formula})'
            else:
                target_cell.value = f'=IF(OR({conds}),"依頼","")'
            written += 1

        # 工具に紐づかない行には自動測定データ反映式のみを設定
        for measure_no, data_index in measure_no_to_data_index.items():
            mr = measure_no_to_row.get(measure_no)
            if mr is None or mr in measure_row_to_tool_rows:
                continue
            target_found += 1
            target_cell = ws.cell(mr, col_idx)
            if not _is_empty_cell_value(target_cell.value):
                continue
            auto_formula = _build_auto_data_formula(
                col_letter=col_letter,
                data_start_row=auto_data_start_row,
                data_index=data_index,
                sep=formula_arg_sep,
            )
            target_cell.value = f"={auto_formula}"
            written += 1

    if target_found == 0:
        raise ValueError(
            "依頼/自動測定データの書き込み対象が1件も見つかりませんでした。\n\n"
            f"- 読み取れた測定No件数: {len(measure_no_to_row)}\n"
            f"- 工具件数: {len(tools)}\n"
            f"- 逆引き対象の測定行件数: {len(measure_row_to_tool_rows)}\n"
            "- 出力列: L～SR（固定）\n"
            f"- 指定Noが見つからない例(先頭10件): {missing_nos[:10]}\n\n"
            "原因候補:\n"
            "1) 測定No列/行範囲が実シートと違う\n"
            "2) 測定Noが数式で、キャッシュ値が未保存（Excelで一度保存してから再実行）\n"
            "3) tool_to_measure_nos のNoがシートに存在しない"
        )

    normalized_count = _normalize_single_cell_array_formulas_in_column(
        ws,
        "B",
        row_start=measure_row_min,
        row_end=measure_row_max,
    )
    if normalized_count:
        print(f"[info] B列の単一セル配列数式を通常数式へ変換: {normalized_count}件")

    _mark_workbook_for_full_recalc(wb)
    saved_path = _save_workbook_atomic(wb, out_path, parent=parent)
    _close_workbook_quietly(wb_values)
    _close_workbook_quietly(wb)
    _restore_package_parts_from_source(xlsx_path, saved_path)
    _force_excel_recalc_and_save(saved_path)
    return saved_path


def _parse_int_list(text: str):
    if not text or not text.strip():
        return []
    parts = re.split(r"[,\s、，;；]+", text.strip())
    nums = []
    for part in parts:
        if not part:
            continue
        value = _try_extract_int(part)
        if value is None:
            raise ValueError(f"測定Noに整数以外の入力が含まれています: '{part}'")
        nums.append(value)
    return nums


def write_measurement_not_required(
    xlsx_path: str,
    out_path: str,
    cfg: dict,
    target_nos: list = None,
    *,
    parent=None,
):
    """
    工具開始行-3行目のE列に"測定不要"を書き込み、指定したNo.の行のL～SR列に条件付きで"-"を書き込む

    Args:
        xlsx_path: 入力Excelファイルのパス
        out_path: 出力Excelファイルのパス
        cfg: 設定辞書（sheet_name, measure_no_col, measure_row_min, measure_row_max を含む）
        target_nos: L～SR列に"-"を書き込む対象のNo.リスト（A列から検索）
        parent: 親ウィンドウ（エラーメッセージ表示用）

    Returns:
        保存されたファイルのパス
    """
    if target_nos is None:
        target_nos = []

    sheet_name = cfg.get("sheet_name", "工程内検査シート")
    measure_no_col = cfg.get("measure_no_col", "A")
    tool_start_row = int(cfg.get("tool_start_row", 200))
    measure_row_min = int(cfg.get("measure_row_min", 11))
    measure_row_step = int(cfg.get("measure_row_step", 3))
    measure_row_max = int(cfg.get("measure_row_max", tool_start_row - 4))

    wb = load_workbook(xlsx_path)
    wb_values = load_workbook(xlsx_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"シート '{sheet_name}' が見つかりません。存在: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    ws_values = wb_values[sheet_name]

    # 1) 測定No(A列の整数) → 行番号のマッピングを作成
    no_col = column_index_from_string(measure_no_col)
    measure_no_to_row = {}
    max_r = min(measure_row_max, ws.max_row or measure_row_max)
    debug_info = []  # デバッグ情報を収集
    
    for r in range(measure_row_min, max_r + 1):
        v = ws_values.cell(r, no_col).value
        cell_obj = ws.cell(r, no_col)
        formula = None
        
        if v is None:
            v = cell_obj.value
            # 数式の場合、数式文字列が返される
            if isinstance(v, str) and v.startswith('='):
                formula = v
        
        if v is None:
            continue

        no = _resolve_measure_no(v, r, measure_row_min, measure_row_step)

        # デバッグ情報を記録（最初の10行のみ）
        if len(debug_info) < 10:
            debug_info.append(
                f"行{r}: 値={v!r}, 型={type(v).__name__}, 数式={formula}, 解決No={no}"
            )

        if no is None:
            continue
        measure_no_to_row[no] = r

    target_row = tool_start_row - 3

    # 2) 指定行のE列に"測定不要"を書き込み（工具開始行-3）
    e_col = column_index_from_string("E")
    target_e_cell = ws.cell(target_row, e_col)
    target_e_cell.value = "測定不要"

    # 3) 指定したNo.の行のL～SR列に条件付きで"-"を書き込む
    # 条件: 指定行（target_row）の同じ列に値が入ると、該当行のL～SR列に"-"が入る
    flag_col_start = column_index_from_string("L")
    flag_col_end = column_index_from_string("SR")

    written_count = 0
    not_found_nos = []
    for no in target_nos:
        row = measure_no_to_row.get(no)
        if row is None:
            # 指定したNo.が見つからない場合は記録
            not_found_nos.append(no)
            continue

        for col_idx in range(flag_col_start, flag_col_end + 1):
            col_letter = get_column_letter(col_idx)
            target_cell = f"{col_letter}{target_row}"
            formula = f'=IF(IFERROR(LEN(TRIM({target_cell}&"")),0)>0,"-","")'
            current_cell = ws.cell(row, col_idx)
            if not _can_overwrite_with_formula(current_cell.value):
                continue
            current_cell.value = formula
        written_count += 1

    if written_count == 0 and target_nos:
        available_nos = sorted(measure_no_to_row.keys())
        available_nos_str = ", ".join(map(str, available_nos[:20]))
        if len(available_nos) > 20:
            available_nos_str += f" ... (他{len(available_nos) - 20}件)"
        
        debug_str = "\n".join(debug_info)
        
        raise ValueError(
            f"指定したNo.の行が見つかりませんでした。\n\n"
            f"【指定したNo.】\n{target_nos}\n\n"
            f"【読み取れた測定No】\n{available_nos_str}\n\n"
            f"【設定】\n"
            f"- 測定No列: {measure_no_col}列\n"
            f"- 測定行範囲: {measure_row_min}〜{measure_row_max}行目\n"
            f"- 読み取れた測定No件数: {len(measure_no_to_row)}件\n\n"
            f"【デバッグ情報（先頭10行）】\n{debug_str}\n\n"
            f"※測定Noが数式の場合、Excelで一度ファイルを開いて保存してから再実行してください。"
        )

    normalized_count = _normalize_single_cell_array_formulas_in_column(
        ws,
        "B",
        row_start=measure_row_min,
        row_end=measure_row_max,
    )
    if normalized_count:
        print(f"[info] B列の単一セル配列数式を通常数式へ変換: {normalized_count}件")

    _mark_workbook_for_full_recalc(wb)
    saved_path = _save_workbook_atomic(wb, out_path, parent=parent)
    _close_workbook_quietly(wb_values)
    _close_workbook_quietly(wb)
    _restore_package_parts_from_source(xlsx_path, saved_path)
    _force_excel_recalc_and_save(saved_path)
    return saved_path


class LoadingDialog(tk.Toplevel):
    """ローディング表示用のダイアログ"""
    def __init__(self, parent, title="処理中...", message="お待ちください..."):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.grab_set()
        
        # ウィンドウを中央に配置
        self.geometry("300x120")
        window_width = 300
        window_height = 120
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        # ウィンドウを閉じるボタンを無効化
        self.protocol("WM_DELETE_WINDOW", lambda: None)
        
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)
        
        ttk.Label(frame, text=message, font=("", 10)).pack(pady=(10, 15))
        
        # プログレスバー
        self.progress = ttk.Progressbar(
            frame,
            mode="indeterminate",
            length=250
        )
        self.progress.pack(pady=10)
        self.progress.start(10)
        
        self.update()

    def close(self):
        """ダイアログを閉じる"""
        self.progress.stop()
        self.grab_release()
        self.destroy()


class ConfigEditor(tb.Window):
    def __init__(self):
        super().__init__(themename="darkly")
        self.title("検査シート 設定作成")
        self.geometry("820x640")
        try:
            self.state("zoomed")
        except Exception:
            pass

        measure_row_min_default = 11
        measure_row_step_default = 3
        tool_start_default = 200
        measure_row_max_default = max(tool_start_default - 4, measure_row_min_default)
        summary_row_min_default = measure_row_min_default
        summary_row_max_default = measure_row_max_default
        summary_row_step_default = measure_row_step_default
        tool_row_step_default = measure_row_step_default

        self.vars = {
            "sheet_name": tk.StringVar(value="工程内検査シート"),
            "measure_no_col": tk.StringVar(value="A"),
            "measure_row_min": tk.IntVar(value=measure_row_min_default),
            "measure_row_max": tk.IntVar(value=measure_row_max_default),
            "measure_row_step": tk.IntVar(value=measure_row_step_default),
            "summary_row_min": tk.IntVar(value=summary_row_min_default),
            "summary_row_max": tk.IntVar(value=summary_row_max_default),
            "summary_row_step": tk.IntVar(value=summary_row_step_default),
            "formula_arg_sep": tk.StringVar(value=","),
            "tool_start_row": tk.IntVar(value=tool_start_default),
            "tool_name_col": tk.StringVar(value="E"),
            "tool_row_step": tk.IntVar(value=tool_row_step_default),
            "auto_data_start_row": tk.IntVar(value=AUTO_DATA_START_ROW_DEFAULT),
            "not_required_row": tk.StringVar(
                value=str(max(tool_start_default - 3, 1))
            ),
            "not_required_nos": tk.StringVar(value=""),
        }

        self.auto_map_measure_no_var = tk.StringVar(value="")
        self.auto_map_data_index_var = tk.StringVar(value="")

        self._bind_basic_setting_sync()

        self.selected_xlsx = tk.StringVar(value="")
        self.preview_title = tk.StringVar(value="プレビュー (未読み込み)")

        self._build_ui()

    def _build_ui(self):
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", padx=10, pady=(10, 0))
        ttk.Button(header_frame, text="ヘルプ", command=self._show_help).pack(side="right")

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.main_canvas = tk.Canvas(body, highlightthickness=0, bd=0)
        self.main_scrollbar = ttk.Scrollbar(
            body,
            orient="vertical",
            command=self.main_canvas.yview,
        )
        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)

        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.main_scrollbar.pack(side="right", fill="y")

        main = ttk.Frame(self.main_canvas, padding=10)
        self.main_canvas_window = self.main_canvas.create_window(
            (0, 0),
            window=main,
            anchor="nw",
        )

        def _update_main_scroll_region(event=None):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

        def _fit_main_width(event):
            self.main_canvas.itemconfigure(self.main_canvas_window, width=event.width)

        main.bind("<Configure>", _update_main_scroll_region)
        self.main_canvas.bind("<Configure>", _fit_main_width)
        self.bind_all("<MouseWheel>", self._on_main_mousewheel, add="+")
        self.bind_all("<Button-4>", self._on_main_mousewheel, add="+")
        self.bind_all("<Button-5>", self._on_main_mousewheel, add="+")

        # 元Excel読み込み＆プレビュー
        source_frame = ttk.LabelFrame(main, text="元Excelとプレビュー", padding=10)
        source_frame.pack(fill="both", expand=True)

        src_row = ttk.Frame(source_frame)
        src_row.pack(fill="x", pady=(0, 8))
        ttk.Label(src_row, text="元Excelファイル").pack(side="left")
        ttk.Entry(src_row, textvariable=self.selected_xlsx, width=60, state="readonly").pack(side="left", padx=6)
        ttk.Button(src_row, text="Excelを選択", command=self._load_preview).pack(side="left")
        ttk.Button(src_row, text="プレビュー更新", command=self._render_preview).pack(side="left", padx=6)
        ttk.Label(src_row, textvariable=self.preview_title).pack(side="right")

        preview_container = ttk.Frame(source_frame)
        preview_container.pack(fill="both", expand=True)
        self.preview_columns = ("A", "B", "G", "K")

        style = ttk.Style(self)
        style.configure(
            "Preview.Treeview",
            rowheight=22,
            borderwidth=1,
            relief="solid",
            bordercolor="#ffffff",
            lightcolor="#ffffff",
            darkcolor="#ffffff",
        )
        style.configure(
            "Preview.Treeview.Heading",
            borderwidth=1,
            relief="solid",
            anchor="center",
        )

        self.preview_tree = ttk.Treeview(
            preview_container,
            columns=self.preview_columns,
            show="headings",
            height=8,
            style="Preview.Treeview",
        )
        vsb = ttk.Scrollbar(preview_container, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=vsb.set)
        self.preview_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")

        basic = ttk.LabelFrame(main, text="基本設定", padding=10)
        basic.pack(fill="x", pady=(10, 0))

        basic_inner = ttk.Frame(basic)
        basic_inner.pack(fill="x")
        basic_inner.columnconfigure(0, weight=1)
        basic_inner.columnconfigure(1, weight=1)

        basic_left = ttk.Frame(basic_inner)
        basic_left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        basic_right = ttk.LabelFrame(basic_inner, text="測定不要書き込み設定", padding=10)
        basic_right.grid(row=0, column=1, sticky="nsew")

        def add_field(row, col, label, key, width=12):
            col_offset = col * 2
            ttk.Label(basic_left, text=label).grid(row=row, column=col_offset, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(basic_left, textvariable=self.vars[key], width=width).grid(row=row, column=col_offset + 1, sticky="w", padx=(0, 20), pady=3)

        add_field(0, 0, "シート名", "sheet_name", width=25)
        add_field(1, 0, "測定No列", "measure_no_col")
        add_field(2, 0, "測定行(min)", "measure_row_min")
        add_field(3, 0, "測定行(max)", "measure_row_max")
        add_field(4, 0, "測定行ステップ", "measure_row_step")
        add_field(5, 0, "集計行(min)", "summary_row_min")
        add_field(6, 0, "集計行(max)", "summary_row_max")
        add_field(7, 0, "集計行ステップ", "summary_row_step")
        add_field(0, 1, "数式区切り(, / ;)", "formula_arg_sep")
        add_field(1, 1, "工具開始行", "tool_start_row")
        add_field(2, 1, "工具名列", "tool_name_col")
        add_field(3, 1, "工具行ステップ", "tool_row_step")
        add_field(4, 1, "自動測定データ開始行", "auto_data_start_row")

        ttk.Label(basic, text="出力列: L～SR（固定）").pack(anchor="w", pady=3)
        ttk.Label(basic_right, text="E列 (工具開始行-3) の行番号:").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=5)
        ttk.Entry(
            basic_right,
            textvariable=self.vars["not_required_row"],
            width=15,
            state="readonly",
        ).grid(row=0, column=1, sticky="w", padx=(0, 30), pady=5)
        ttk.Label(
            basic_right,
            text="L～SR列に'-'を入れるNo.(カンマ区切り):",
        ).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=5)
        ttk.Entry(basic_right, textvariable=self.vars["not_required_nos"], width=40).grid(row=1, column=1, sticky="ew", padx=(0, 10), pady=5)
        basic_right.columnconfigure(1, weight=1)

        tools_frame = ttk.LabelFrame(main, text="工具と測定No対応", padding=10)
        tools_frame.pack(fill="both", expand=True, pady=(10, 0))
        self.tools_tree = ttk.Treeview(tools_frame, columns=("tool", "nos"), show="headings", height=12)
        self.tools_tree.heading("tool", text="工具名")
        self.tools_tree.heading("nos", text="測定No(カンマ区切り)")
        self.tools_tree.column("tool", width=220, anchor="w")
        self.tools_tree.column("nos", width=420, anchor="w")
        self.tools_tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tools_frame, orient="vertical", command=self.tools_tree.yview)
        self.tools_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        tools_btns = ttk.Frame(main)
        tools_btns.pack(fill="x", pady=(6, 0))
        ttk.Button(tools_btns, text="工具追加", command=self._add_tool_dialog).pack(side="left")
        ttk.Button(tools_btns, text="選択編集", command=self._edit_selected_tool).pack(side="left", padx=5)
        ttk.Button(tools_btns, text="選択削除", command=self._delete_selected_tool).pack(side="left")

        auto_map_frame = ttk.LabelFrame(main, text="自動測定データ対応（測定No → データ順番）", padding=10)
        auto_map_frame.pack(fill="both", expand=True, pady=(10, 0))

        auto_input = ttk.Frame(auto_map_frame)
        auto_input.pack(fill="x", pady=(0, 8))
        ttk.Label(auto_input, text="測定No").pack(side="left")
        ttk.Entry(auto_input, textvariable=self.auto_map_measure_no_var, width=12).pack(
            side="left", padx=(6, 12)
        )
        ttk.Label(auto_input, text=f"データ順番 (1〜{AUTO_DATA_MAX_ITEMS})").pack(side="left")
        ttk.Entry(auto_input, textvariable=self.auto_map_data_index_var, width=12).pack(
            side="left", padx=(6, 12)
        )
        ttk.Button(auto_input, text="追加", command=self._add_auto_map).pack(side="left")
        ttk.Button(auto_input, text="選択削除", command=self._delete_selected_auto_map).pack(
            side="left", padx=6
        )

        self.auto_map_tree = ttk.Treeview(
            auto_map_frame,
            columns=("measure_no", "data_index"),
            show="headings",
            height=6,
        )
        self.auto_map_tree.heading("measure_no", text="測定No")
        self.auto_map_tree.heading("data_index", text="データ順番")
        self.auto_map_tree.column("measure_no", width=200, anchor="w")
        self.auto_map_tree.column("data_index", width=140, anchor="center")
        self.auto_map_tree.pack(side="left", fill="both", expand=True)
        auto_scrollbar = ttk.Scrollbar(
            auto_map_frame,
            orient="vertical",
            command=self.auto_map_tree.yview,
        )
        self.auto_map_tree.configure(yscrollcommand=auto_scrollbar.set)
        auto_scrollbar.pack(side="right", fill="y")

        btns = ttk.Frame(main)
        btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="この設定でExcel生成", command=self._run_build).pack(side="right")

        if not self.tools_tree.get_children():
            self._insert_tool("前挽き(サンプル)", "1, 5, 10")

    def _is_in_main_content(self, widget):
        current = widget
        while current is not None:
            if current is self.main_canvas:
                return True
            parent_name = current.winfo_parent()
            if not parent_name:
                break
            try:
                current = current.nametowidget(parent_name)
            except Exception:
                break
        return False

    def _on_main_mousewheel(self, event):
        if not self._is_in_main_content(event.widget):
            return
        if getattr(event, "delta", 0):
            self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif getattr(event, "num", None) == 4:
            self.main_canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.main_canvas.yview_scroll(1, "units")

    def _bind_basic_setting_sync(self):
        self.vars["tool_start_row"].trace_add("write", self._sync_tool_start_row)
        self.vars["measure_row_min"].trace_add("write", self._sync_measure_row_min)
        self.vars["measure_row_max"].trace_add("write", self._sync_measure_row_max)
        self.vars["measure_row_step"].trace_add("write", self._sync_measure_row_step)

    def _sync_tool_start_row(self, *args):
        try:
            tool_start = self.vars["tool_start_row"].get()
        except tk.TclError:
            return
        min_row = self.vars["measure_row_min"].get()
        desired_max = max(tool_start - 4, min_row)
        if self.vars["measure_row_max"].get() != desired_max:
            self.vars["measure_row_max"].set(desired_max)
        target_row = max(tool_start - 3, 1)
        desired_row_str = str(target_row)
        if self.vars["not_required_row"].get() != desired_row_str:
            self.vars["not_required_row"].set(desired_row_str)

    def _sync_measure_row_min(self, *args):
        try:
            min_row = self.vars["measure_row_min"].get()
        except tk.TclError:
            return
        self.vars["summary_row_min"].set(min_row)
        if self.vars["measure_row_max"].get() < min_row:
            self.vars["measure_row_max"].set(min_row)

    def _sync_measure_row_max(self, *args):
        try:
            max_row = self.vars["measure_row_max"].get()
        except tk.TclError:
            return
        self.vars["summary_row_max"].set(max_row)

    def _sync_measure_row_step(self, *args):
        try:
            step = self.vars["measure_row_step"].get()
        except tk.TclError:
            return
        if step < 1:
            self.vars["measure_row_step"].set(1)
            return
        self.vars["summary_row_step"].set(step)
        if self.vars["tool_row_step"].get() != step:
            self.vars["tool_row_step"].set(step)

    def _load_preview(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="元の検査シート（xlsx）を選択",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        self.selected_xlsx.set(path)
        
        # ローディングダイアログを表示して別スレッドでプレビュー読み込み
        loading = LoadingDialog(self, "読み込み中...", "Excelファイルを読み込んでいます...")
        
        result = {"success": False, "error": None}
        
        def load_task():
            try:
                self._render_preview_internal()
                result["success"] = True
            except Exception as e:
                result["error"] = e
        
        thread = threading.Thread(target=load_task, daemon=True)
        thread.start()
        
        # スレッドが完了するまで待機
        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
            else:
                loading.close()
                if result["success"]:
                    pass  # 成功時は何もしない
                elif result["error"]:
                    messagebox.showerror("読み込み失敗", str(result["error"]), parent=self)
        
        self.after(100, check_completion)

    def _render_preview(self):
        """プレビュー更新ボタン用のラッパー"""
        path = self.selected_xlsx.get().strip()
        if not path:
            messagebox.showinfo("読み込み待ち", "先にExcelファイルを選択してください。", parent=self)
            return
        
        loading = LoadingDialog(self, "更新中...", "プレビューを更新しています...")
        
        result = {"success": False, "error": None}
        
        def render_task():
            try:
                self._render_preview_internal()
                result["success"] = True
            except Exception as e:
                result["error"] = e
        
        thread = threading.Thread(target=render_task, daemon=True)
        thread.start()
        
        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
            else:
                loading.close()
                if result["success"]:
                    pass
                elif result["error"]:
                    messagebox.showerror("更新失敗", str(result["error"]), parent=self)
        
        self.after(100, check_completion)

    def _render_preview_internal(self):
        """実際のプレビュー処理（従来の_render_previewの内容）"""
        path = self.selected_xlsx.get().strip()
        if not path:
            return

        sheet_name = self.vars["sheet_name"].get().strip() or "工程内検査シート"
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            raise Exception(f"Excelを開けませんでした。\n{e}")

        if sheet_name not in wb.sheetnames:
            raise Exception(
                f"シート「{sheet_name}」が見つかりません。\nシート名を確認して再度プレビューしてください。"
            )

        ws = wb[sheet_name]
        preview_col_indices = [column_index_from_string(c) for c in self.preview_columns]

        if not ws.max_row or not ws.max_column:
            raise Exception("表示できるデータがありません。")

        col_widths = {"A": 80, "B": 120, "G": 140, "K": 140}
        
        # UIの更新はメインスレッドで実行
        def update_ui():
            self.preview_tree.configure(columns=self.preview_columns)
            for col in self.preview_columns:
                self.preview_tree.heading(col, text=col, anchor="center")
                self.preview_tree.column(
                    col,
                    width=col_widths.get(col, 120),
                    minwidth=60,
                    anchor="center",
                    stretch=False,
                )

            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)

            header_row = 10
            group_size = 3
            max_row_value = ws.max_row or 0

            if max_row_value < header_row:
                raise Exception(f"{header_row}行目以降に表示できるデータがありません。")

            # ヘッダー行(10行目)を挿入
            header_vals = []
            for c in preview_col_indices:
                v = ws.cell(header_row, c).value
                header_vals.append("" if v is None else str(v))
            self.preview_tree.insert("", "end", values=header_vals)

            data_start_row = header_row + 1
            row_count = 0
            for r in range(data_start_row, max_row_value + 1, group_size):
                a_val = ws.cell(r, preview_col_indices[0]).value
                if a_val is None or (isinstance(a_val, str) and not a_val.strip()):
                    break

                row_vals = []
                for c in preview_col_indices:
                    v = ws.cell(r, c).value
                    row_vals.append("" if v is None else str(v))

                self.preview_tree.insert("", "end", values=row_vals)
                row_count += 1

            if row_count == 0:
                for item in self.preview_tree.get_children():
                    self.preview_tree.delete(item)
                raise Exception("11行目以降の先頭行(A列)が空のため表示できるデータがありません。")

            self.preview_title.set(
                f"{sheet_name} プレビュー（{row_count}行: 10行目ヘッダー＋11行目以降3行刻み先頭行）"
            )
        
        self.after(0, update_ui)

    def _insert_tool(self, tool_name: str, nos_text: str):
        self.tools_tree.insert("", "end", values=(tool_name, nos_text))

    def _tool_dialog(self, title, init_tool="", init_nos=""):
        win = tk.Toplevel(self)
        win.title(title)
        win.transient(self)
        win.grab_set()

        tool_var = tk.StringVar(value=init_tool)
        nos_var = tk.StringVar(value=init_nos)

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="工具名").grid(row=0, column=0, sticky="w", pady=4)
        tool_entry = ttk.Entry(frm, textvariable=tool_var, width=30)
        tool_entry.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="測定No(カンマ区切り)").grid(
            row=1, column=0, sticky="w", pady=4
        )
        ttk.Entry(frm, textvariable=nos_var, width=40).grid(
            row=1, column=1, sticky="w", pady=4
        )

        result = {"ok": False}

        def on_ok():
            tool = tool_var.get().strip()
            if not tool:
                messagebox.showwarning(
                    "入力不足", "工具名を入力してください。", parent=win
                )
                return
            try:
                _parse_int_list(nos_var.get())
            except Exception:
                messagebox.showwarning(
                    "入力エラー",
                    "測定Noは整数のカンマ区切りで入力してください。",
                    parent=win,
                )
                return
            result["ok"] = True
            result["tool"] = tool
            result["nos"] = nos_var.get().strip()
            win.destroy()

        def on_cancel():
            win.destroy()

        bfrm = ttk.Frame(frm)
        bfrm.grid(row=2, column=0, columnspan=2, sticky="e", pady=(8, 0))
        ttk.Button(bfrm, text="OK", command=on_ok).pack(side="right")
        ttk.Button(bfrm, text="キャンセル", command=on_cancel).pack(
            side="right", padx=5
        )

        tool_entry.focus_set()
        self.wait_window(win)
        return result

    def _add_tool_dialog(self):
        res = self._tool_dialog("工具追加")
        if res.get("ok"):
            self._insert_tool(res["tool"], res["nos"])

    def _edit_selected_tool(self):
        sel = self.tools_tree.selection()
        if not sel:
            messagebox.showinfo(
                "選択なし", "編集する行を選択してください。", parent=self
            )
            return
        item = sel[0]
        tool, nos = self.tools_tree.item(item, "values")
        res = self._tool_dialog("工具編集", init_tool=tool, init_nos=nos)
        if res.get("ok"):
            self.tools_tree.item(item, values=(res["tool"], res["nos"]))

    def _delete_selected_tool(self):
        sel = self.tools_tree.selection()
        if not sel:
            return
        if not messagebox.askyesno(
            "削除確認", "選択した工具を削除しますか？", parent=self
        ):
            return
        for item in sel:
            self.tools_tree.delete(item)

    def _add_auto_map(self):
        measure_no_raw = self.auto_map_measure_no_var.get().strip()
        data_index_raw = self.auto_map_data_index_var.get().strip()

        key = _normalize_measure_no_key(measure_no_raw)
        data_index = _try_extract_int(data_index_raw)

        if key == "":
            messagebox.showwarning("入力エラー", "測定Noを入力してください。", parent=self)
            return
        if data_index is None or data_index < 1 or data_index > AUTO_DATA_MAX_ITEMS:
            messagebox.showwarning(
                "入力エラー",
                f"データ順番は1〜{AUTO_DATA_MAX_ITEMS}の整数で入力してください。",
                parent=self,
            )
            return

        key_str = str(key)
        existing = None
        for item in self.auto_map_tree.get_children():
            values = self.auto_map_tree.item(item, "values")
            if str(values[0]) == key_str:
                existing = item
                break

        if existing is not None:
            self.auto_map_tree.item(existing, values=(key_str, str(data_index)))
        else:
            self.auto_map_tree.insert("", "end", values=(key_str, str(data_index)))

        self.auto_map_measure_no_var.set("")
        self.auto_map_data_index_var.set("")

    def _delete_selected_auto_map(self):
        sel = self.auto_map_tree.selection()
        if not sel:
            return
        for item in sel:
            self.auto_map_tree.delete(item)

    def _gather_cfg(self):
        try:
            tools = []
            tool_to_measure_nos = {}
            measure_no_to_data_index = {}
            for item in self.tools_tree.get_children():
                tool, nos_text = self.tools_tree.item(item, "values")
                tools.append(tool)
                tool_to_measure_nos[tool] = _parse_int_list(nos_text)

            for item in self.auto_map_tree.get_children():
                measure_no_text, data_index_text = self.auto_map_tree.item(item, "values")
                key = _normalize_measure_no_key(measure_no_text)
                data_index = _try_extract_int(data_index_text)
                if key == "" or data_index is None:
                    continue
                measure_no_to_data_index[key] = data_index

            if not tools:
                raise ValueError("工具が1件もありません。")

            cfg = {
                "sheet_name": self.vars["sheet_name"].get().strip(),
                "measure_no_col": self.vars["measure_no_col"].get().strip().upper(),
                "measure_row_min": int(self.vars["measure_row_min"].get()),
                "measure_row_max": int(self.vars["measure_row_max"].get()),
                "measure_row_step": int(self.vars["measure_row_step"].get()),
                "summary_row_min": int(self.vars["summary_row_min"].get()),
                "summary_row_max": int(self.vars["summary_row_max"].get()),
                "summary_row_step": int(self.vars["summary_row_step"].get()),
                "formula_arg_sep": self.vars["formula_arg_sep"].get().strip(),
                "tool_start_row": int(self.vars["tool_start_row"].get()),
                "tool_name_col": self.vars["tool_name_col"].get().strip().upper(),
                "tool_row_step": int(self.vars["tool_row_step"].get()),
                "auto_data_start_row": int(self.vars["auto_data_start_row"].get()),
                "measure_no_to_data_index": measure_no_to_data_index,
                "tools": tools,
                "tool_to_measure_nos": tool_to_measure_nos,
            }
            if not cfg["sheet_name"]:
                raise ValueError("シート名が空です。")
            return cfg
        except Exception as e:
            raise ValueError(f"設定の取得に失敗: {e}")

    def _run_build(self):
        try:
            cfg = self._gather_cfg()
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return

        xlsx = self.selected_xlsx.get().strip()
        if not xlsx:
            messagebox.showinfo(
                "ファイル未選択",
                "先に「Excelを選択」で元ファイルを読み込んでください。",
                parent=self,
            )
            self._load_preview()
            xlsx = self.selected_xlsx.get().strip()
            if not xlsx:
                return
        out_path = pick_save_path(
            "出力先（生成したxlsx）を保存",
            ".xlsx",
            [("Excel", "*.xlsx")],
            parent=self,
        )
        if not out_path:
            return

        # ローディングダイアログを表示して別スレッドで生成処理
        loading = LoadingDialog(self, "生成中...", "Excelファイルを生成しています...")
        
        result = {"success": False, "error": None, "saved_path": None}
        
        def build_task():
            try:
                saved_path = build_request_formulas(xlsx, out_path, cfg, parent=self)
                result["saved_path"] = saved_path
                
                # 測定不要書き込み設定が入力されている場合は実行
                not_required_nos_text = self.vars["not_required_nos"].get().strip()
                if not_required_nos_text:
                    try:
                        target_nos = _parse_int_list(not_required_nos_text)
                        if target_nos:
                            saved_path = write_measurement_not_required(
                                saved_path,
                                out_path,
                                cfg,
                                target_nos=target_nos,
                                parent=self,
                            )
                            result["saved_path"] = saved_path
                    except Exception as e:
                        result["error"] = f"測定不要書き込みでエラーが発生しました:\n{e}"
                        result["success"] = True  # メイン処理は成功
                        return
                
                result["success"] = True
            except Exception as e:
                result["error"] = str(e)
        
        thread = threading.Thread(target=build_task, daemon=True)
        thread.start()
        
        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
            else:
                loading.close()
                if result["success"]:
                    if result["error"]:
                        # 警告メッセージ
                        messagebox.showwarning(
                            "警告",
                            f"Excel生成は完了しましたが、{result['error']}",
                            parent=self,
                        )
                    else:
                        messagebox.showinfo("完了", f"生成しました:\n{result['saved_path']}", parent=self)
                else:
                    messagebox.showerror("失敗", result["error"], parent=self)
        
        self.after(100, check_completion)

    def _run_write_not_required(self):
        """測定不要書き込み機能の実行"""
        try:
            cfg = self._gather_cfg()
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return

        # UI内の入力フィールドから値を取得
        try:
            target_nos = _parse_int_list(self.vars["not_required_nos"].get())
        except Exception as e:
            messagebox.showerror(
                "入力エラー",
                f"測定Noの入力が不正です。\n{e}",
                parent=self,
            )
            return

        if not target_nos:
            messagebox.showwarning(
                "入力不足",
                "L～SR列に'-'を入れるNo.を入力してください。",
                parent=self,
            )
            return

        xlsx = self.selected_xlsx.get().strip()
        if not xlsx:
            messagebox.showinfo(
                "ファイル未選択",
                "先に「Excelを選択」で元ファイルを読み込んでください。",
                parent=self,
            )
            self._load_preview()
            xlsx = self.selected_xlsx.get().strip()
            if not xlsx:
                return
        out_path = pick_save_path(
            "出力先（生成したxlsx）を保存",
            ".xlsx",
            [("Excel", "*.xlsx")],
            parent=self,
        )
        if not out_path:
            return

        try:
            saved_path = write_measurement_not_required(
                xlsx,
                out_path,
                cfg,
                target_nos=target_nos,
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return
        messagebox.showinfo("完了", f"書き込み完了:\n{saved_path}", parent=self)

    def _show_help(self):
        """ヘルプモーダルを表示"""
        help_window = tk.Toplevel(self)
        help_window.title("使い方")
        help_window.transient(self)
        help_window.grab_set()
        help_window.geometry("600x500")

        # スクロール可能なフレーム
        canvas = tk.Canvas(help_window, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(help_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # コンテンツ
        content_frame = ttk.Frame(scrollable_frame, padding=20)
        content_frame.pack(fill="both", expand=True)

        # バージョン情報
        version_frame = ttk.LabelFrame(content_frame, text="バージョン情報", padding=10)
        version_frame.pack(fill="x", pady=(0, 15))

        ttk.Label(version_frame, text="Version: 0.6.0", font=("", 10, "bold")).pack(
            anchor="w"
        )
        ttk.Label(version_frame, text="Latest: 2025-12-15", font=("", 10)).pack(
            anchor="w", pady=(5, 0)
        )
        ttk.Label(
            version_frame, text="created by DIP Dpertment/A・T", font=("", 10)
        ).pack(anchor="w", pady=(5, 0))

        # 使い方
        usage_frame = ttk.LabelFrame(content_frame, text="使い方", padding=10)
        usage_frame.pack(fill="both", expand=True)

        usage_text = """
【基本設定】
1. シート名、測定No列、行範囲などを設定します
2. デフォルト値が設定されているので、必要に応じて変更してください

【工具と測定No対応】
1. 「工具追加」ボタンで工具を追加します
2. 工具名と対応する測定Noをカンマ区切りで入力します
  例: 1, 5, 10, 15
3. 「選択編集」「選択削除」で既存の工具を編集・削除できます

【測定不要書き込み設定】
1. 工具開始行-3行目のE列に「測定不要」が自動で書き込まれます
2. L～SR列に「-」を入れる測定Noをカンマ区切りで指定します
3. この設定は任意です（空欄の場合はスキップされます）

【自動測定データ対応】
1. 「自動測定データ開始行」は既定で230行です
2. 「測定No → データ順番」を追加すると、L～SR列で同一列の下部データを参照します
3. 工具影響があるセルは「依頼」を優先し、それ以外は測定データ参照式を設定します

【Excel生成】
1. 「この設定でExcel生成」ボタンをクリックします
2. 元のExcelファイルを選択します
3. 出力先のファイル名を指定します
4. 処理が完了すると、指定した場所にExcelファイルが生成されます

【生成される内容】
- 指定した工具に対応する測定Noの行に「依頼」という文字が自動入力されます
- L列からSR列までの各列に対して処理が実行されます
- 測定不要設定が指定されている場合は、該当する行に「測定不要」や「-」が書き込まれます
        """

        usage_label = ttk.Label(
            usage_frame,
            text=usage_text.strip(),
            justify="left",
            font=("", 9),
        )
        usage_label.pack(anchor="w", padx=5)

        # 注意事項
        note_frame = ttk.LabelFrame(content_frame, text="注意事項", padding=10)
        note_frame.pack(fill="x", pady=(15, 0))

        note_text = """
• 元のExcelファイルは開いていない状態で実行してください
• 出力先ファイルが既に存在する場合は上書きされます
• 測定Noは整数で指定してください
• シート名が存在しない場合はエラーになります
        """

        note_label = ttk.Label(
            note_frame,
            text=note_text.strip(),
            justify="left",
            font=("", 9),
            foreground="firebrick",
        )
        note_label.pack(anchor="w", padx=5)

        # 閉じるボタン
        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(fill="x", pady=(15, 0))
        ttk.Button(btn_frame, text="閉じる", command=help_window.destroy).pack(
            side="right"
        )

        # スクロールバーとキャンバスの配置
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # マウスホイールでスクロール（Windows用）
        def _on_mousewheel(event):
            if event.delta:
                # Windows
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            else:
                # Linux
                if event.num == 4:
                    canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    canvas.yview_scroll(1, "units")

        # Windows用
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # Linux用
        canvas.bind_all("<Button-4>", _on_mousewheel)
        canvas.bind_all("<Button-5>", _on_mousewheel)

        # キャンバスのスクロール領域を更新
        def update_scroll_region(event=None):
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        def configure_canvas_width(event):
            canvas_width = event.width
            canvas.itemconfig(canvas.find_all()[0], width=canvas_width)

        scrollable_frame.bind("<Configure>", update_scroll_region)
        canvas.bind("<Configure>", configure_canvas_width)

        help_window.focus_set()


def main():
    app = ConfigEditor()
    app.mainloop()


if __name__ == "__main__":
    main()
