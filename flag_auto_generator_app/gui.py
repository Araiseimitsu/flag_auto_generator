import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import ttkbootstrap as tb
from ttkbootstrap.constants import INFO, SECONDARY
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .help_dialog import open_help_window
from .excel_ops import (
    AUTO_DATA_MAX_ITEMS,
    LOCKED_BASIC_SETTINGS,
    NOT_REQUIRED_ROW_DEFAULT,
    _derive_auto_data_start_row,
    _derive_layout_rows,
    _normalize_measure_no_key,
    _parse_int_list,
    _try_extract_int,
    build_request_formulas,
    write_measurement_not_required,
)
from .ui_helpers import LoadingDialog, pick_save_path
from .ui_theme import (
    HINT_WRAPLENGTH,
    THEME_NAME,
    apply_app_style,
    apply_preview_treeview_style,
    font_ui,
    make_step_caption,
    place_toplevel_center,
    scrollstrip_background,
    section_separator,
)


class ConfigEditor(tb.Window):
    def __init__(self):
        super().__init__(themename=THEME_NAME)
        self.title("検査シート 設定ツール")
        self.geometry("920x780")
        self.minsize(780, 600)
        try:
            self.state("zoomed")
        except Exception:
            pass
        try:
            apply_app_style(self)
        except (tk.TclError, AttributeError):
            pass

        measure_row_min_default = LOCKED_BASIC_SETTINGS["measure_row_min"]
        measure_row_step_default = LOCKED_BASIC_SETTINGS["measure_row_step"]
        not_required_row_default = NOT_REQUIRED_ROW_DEFAULT
        self.not_required_row_choices = tuple(
            str(row) for row in range(measure_row_min_default + 3, 302, measure_row_step_default)
        )
        measure_row_max_default, tool_start_default = _derive_layout_rows(
            not_required_row_default,
            measure_row_min_default,
        )
        summary_row_min_default = LOCKED_BASIC_SETTINGS["summary_row_min"]
        summary_row_max_default = measure_row_max_default
        summary_row_step_default = LOCKED_BASIC_SETTINGS["summary_row_step"]
        tool_row_step_default = LOCKED_BASIC_SETTINGS["tool_row_step"]

        self.vars = {
            "sheet_name": tk.StringVar(value="工程内検査シート"),
            "measure_no_col": tk.StringVar(value=LOCKED_BASIC_SETTINGS["measure_no_col"]),
            "measure_row_min": tk.IntVar(value=measure_row_min_default),
            "measure_row_max": tk.IntVar(value=measure_row_max_default),
            "measure_row_step": tk.IntVar(value=measure_row_step_default),
            "summary_row_min": tk.IntVar(value=summary_row_min_default),
            "summary_row_max": tk.IntVar(value=summary_row_max_default),
            "summary_row_step": tk.IntVar(value=summary_row_step_default),
            "formula_arg_sep": tk.StringVar(value=LOCKED_BASIC_SETTINGS["formula_arg_sep"]),
            "tool_start_row": tk.IntVar(value=tool_start_default),
            "tool_name_col": tk.StringVar(value=LOCKED_BASIC_SETTINGS["tool_name_col"]),
            "tool_row_step": tk.IntVar(value=tool_row_step_default),
            "not_required_row": tk.StringVar(value=str(not_required_row_default)),
        }

        self.not_required_no_input_var = tk.StringVar(value="")
        self.auto_map_measure_no_var = tk.StringVar(value="")
        self.auto_map_data_index_var = tk.StringVar(value="")

        self._bind_basic_setting_sync()
        self._apply_locked_basic_settings()

        self.selected_xlsx = tk.StringVar(value="")
        self.preview_title = tk.StringVar(value="まだ表を表示していません")

        self._build_ui()

    def _build_ui(self):
        app_bar = ttk.Frame(self, style="Toolbar.TFrame", padding=(28, 22, 28, 14))
        app_bar.pack(fill=tk.X)
        left = ttk.Frame(app_bar, style="Toolbar.TFrame")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(left, text="FLAG AUTO GENERATOR", style="HeroEyebrow.TLabel").pack(anchor="w")
        ttk.Label(left, text="検査シート 設定", style="HeroTitle.TLabel").pack(anchor="w", pady=(4, 0))
        tb.Button(
            app_bar,
            text="使い方と注意",
            command=self._show_help,
            bootstyle="outline-secondary",
        ).pack(side=tk.RIGHT, pady=(4, 0))

        body = ttk.Frame(self, style="Toolbar.TFrame")
        body.pack(fill=tk.BOTH, expand=True, padx=0, pady=(0, 8))

        scroll_bg = scrollstrip_background()
        self.main_canvas = tk.Canvas(
            body,
            highlightthickness=0,
            bd=0,
            background=scroll_bg,
        )
        self.main_scrollbar = ttk.Scrollbar(
            body,
            orient="vertical",
            command=self.main_canvas.yview,
        )
        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        main = ttk.Frame(self.main_canvas, style="Toolbar.TFrame", padding=(28, 10, 28, 36))
        self.main_canvas_window = self.main_canvas.create_window(
            (0, 0),
            window=main,
            anchor="nw",
        )

        def _update_main_scroll_region(event=None):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

        def _fit_main_width(event):
            self.main_canvas.itemconfigure(self.main_canvas_window, width=event.width)

        main.bind("<Configure>", _update_main_scroll_region)
        self.main_canvas.bind("<Configure>", _fit_main_width)
        self.bind_all("<MouseWheel>", self._on_main_mousewheel, add="+")
        self.bind_all("<Button-4>", self._on_main_mousewheel, add="+")
        self.bind_all("<Button-5>", self._on_main_mousewheel, add="+")

        make_step_caption(
            main,
            1,
            "元の Excel を指定する",
            "会社の雛形ファイル（.xlsx）を選び、下の表で中身の一部を確認します。",
        )
        source_frame = ttk.LabelFrame(
            main,
            text="ファイルと取り込み内容のイメージ",
            style="AppCard.TLabelframe",
            padding=14,
        )
        source_frame.pack(fill=tk.BOTH, expand=True)

        src_file = ttk.Frame(source_frame, style="Surface.TFrame")
        src_file.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(src_file, text="参照中のファイル", style="CardTitle.TLabel").pack(anchor=tk.W)
        path_row = ttk.Frame(source_frame, style="Surface.TFrame")
        path_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Entry(
            path_row,
            textvariable=self.selected_xlsx,
            state="readonly",
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        tb.Button(path_row, text="参照…", command=self._load_preview, bootstyle=SECONDARY).pack(
            side=tk.LEFT
        )
        tb.Button(
            path_row,
            text="表を再表示",
            command=self._render_preview,
            bootstyle="outline-secondary",
        ).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Label(
            source_frame,
            textvariable=self.preview_title,
            style="CardNote.TLabel",
        ).pack(anchor="w", pady=(0, 6))

        preview_container = ttk.Frame(source_frame, style="Surface.TFrame")
        preview_container.pack(fill=tk.BOTH, expand=True)
        self.preview_columns = ("A", "B", "G", "K")

        apply_preview_treeview_style(self)
        ttk.Style(self).configure("Preview.Treeview.Heading", anchor="center")

        self.preview_tree = ttk.Treeview(
            preview_container,
            columns=self.preview_columns,
            show="headings",
            height=9,
            style="Preview.Treeview",
        )
        vsb = ttk.Scrollbar(preview_container, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=vsb.set)
        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        section_separator(main)

        make_step_caption(
            main,
            2,
            "基本のシート名と「測定不要」の行",
            "いつも使うシート名に合わせます。E 列の「測定不要」が入る行を指定します。",
        )
        basic = ttk.LabelFrame(main, text="基本・測定不要", padding=14)
        basic.configure(style="AppCard.TLabelframe")
        basic.pack(fill=tk.X)

        basic_inner = ttk.Frame(basic, style="Surface.TFrame")
        basic_inner.pack(fill=tk.X)
        basic_inner.columnconfigure(0, weight=1)
        basic_inner.columnconfigure(1, weight=1)

        basic_left = ttk.LabelFrame(
            basic_inner,
            text="シート名",
            style="AppCard.TLabelframe",
            padding=10,
        )
        basic_left.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 10))
        ttk.Label(basic_left, text="加工・検査で開くシートの名前", style="CardNote.TLabel").pack(
            anchor=tk.W
        )
        ttk.Entry(
            basic_left,
            textvariable=self.vars["sheet_name"],
            width=28,
        ).pack(anchor=tk.W, pady=(6, 0))

        basic_right = ttk.LabelFrame(
            basic_inner,
            text="測定不要の書き込み（任意）",
            style="AppCard.TLabelframe",
            padding=10,
        )
        basic_right.grid(row=0, column=1, sticky=tk.NSEW)
        ttk.Label(
            basic_right,
            text="E 列の「測定不要」行と、L～SR で「-」にしたい測定 No を設定します。",
        ).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 4))

        not_req_setting_row = ttk.Frame(basic_right, style="Surface.TFrame")
        not_req_setting_row.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(0, 8))
        ttk.Label(not_req_setting_row, text="測定不要の開始行").pack(side=tk.LEFT)
        ttk.Combobox(
            not_req_setting_row,
            textvariable=self.vars["not_required_row"],
            values=self.not_required_row_choices,
            width=10,
            state="readonly",
        ).pack(side=tk.LEFT, padx=(8, 0))

        not_req_row = ttk.Frame(basic_right, style="Surface.TFrame")
        not_req_row.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        ttk.Label(not_req_row, text="測定 No").pack(side=tk.LEFT)
        ttk.Entry(
            not_req_row,
            textvariable=self.not_required_no_input_var,
            width=10,
        ).pack(side=tk.LEFT, padx=(4, 8))
        tb.Button(
            not_req_row,
            text="追加",
            command=self._add_not_required_no,
            bootstyle=SECONDARY,
        ).pack(side=tk.LEFT, padx=(0, 4))
        tb.Button(
            not_req_row,
            text="選択を削除",
            command=self._delete_selected_not_required_no,
            bootstyle="outline-secondary",
        ).pack(side=tk.LEFT)

        nr_list_frame = ttk.Frame(basic_right, style="Surface.TFrame")
        nr_list_frame.grid(row=3, column=0, columnspan=2, sticky=tk.NSEW, pady=(0, 4))
        self.not_required_nos_tree = ttk.Treeview(
            nr_list_frame,
            columns=("no",),
            show="headings",
            height=5,
            style="Data.Treeview",
        )
        self.not_required_nos_tree.heading("no", text="登録した測定 No")
        self.not_required_nos_tree.column("no", width=140, anchor=tk.CENTER)
        nr_sb = ttk.Scrollbar(
            nr_list_frame,
            orient=tk.VERTICAL,
            command=self.not_required_nos_tree.yview,
        )
        self.not_required_nos_tree.configure(yscrollcommand=nr_sb.set)
        self.not_required_nos_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        nr_sb.pack(side=tk.RIGHT, fill=tk.Y)
        basic_right.columnconfigure(0, weight=1)
        section_separator(main)

        make_step_caption(
            main,
            3,
            "工具名と、どの測定番号に当てるか",
            "工具名と、その工具が扱う測定 No を一覧で入れます。行を選ぶと「編集」「削除」ができます。",
        )
        tools_frame = ttk.LabelFrame(
            main,
            text="工具の一覧",
            style="AppCard.TLabelframe",
            padding=10,
        )
        tools_frame.pack(fill=tk.BOTH, expand=True)
        self.tools_tree = ttk.Treeview(
            tools_frame,
            columns=("tool", "nos"),
            show="headings",
            height=11,
            style="Data.Treeview",
        )
        self.tools_tree.heading("tool", text="工具名")
        self.tools_tree.heading("nos", text="測定 No（半角数字・カンマ区切り）")
        self.tools_tree.column("tool", width=220, anchor=tk.W)
        self.tools_tree.column("nos", width=450, anchor=tk.W)
        self.tools_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tools_frame, orient="vertical", command=self.tools_tree.yview)
        self.tools_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        tools_btns = ttk.Frame(main, style="Toolbar.TFrame")
        tools_btns.pack(fill=tk.X, pady=(8, 0))
        tb.Button(
            tools_btns,
            text="＋ 追加",
            command=self._add_tool_dialog,
            bootstyle=SECONDARY,
        ).pack(side=tk.LEFT)
        tb.Button(
            tools_btns,
            text="編集",
            command=self._edit_selected_tool,
            bootstyle="outline-secondary",
        ).pack(side=tk.LEFT, padx=6)
        tb.Button(
            tools_btns,
            text="削除",
            command=self._delete_selected_tool,
            bootstyle="outline-secondary",
        ).pack(side=tk.LEFT)

        section_separator(main)

        make_step_caption(
            main,
            4,
            "自動測定のデータと測定番号の対応",
            f"測定 No ごとに、自動測定ブロック内の何番目の値を参照するか（1〜{AUTO_DATA_MAX_ITEMS}）を入れます。",
        )
        auto_map_frame = ttk.LabelFrame(
            main,
            text="対応表",
            style="AppCard.TLabelframe",
            padding=10,
        )
        auto_map_frame.pack(fill=tk.BOTH, expand=True)

        auto_input = ttk.Frame(auto_map_frame, style="Surface.TFrame")
        auto_input.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(auto_input, text="測定 No").pack(side=tk.LEFT)
        ttk.Entry(
            auto_input,
            textvariable=self.auto_map_measure_no_var,
            width=10,
        ).pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(auto_input, text=f"データの順番（1〜{AUTO_DATA_MAX_ITEMS}）").pack(side=tk.LEFT)
        ttk.Entry(
            auto_input,
            textvariable=self.auto_map_data_index_var,
            width=8,
        ).pack(side=tk.LEFT, padx=(4, 8))
        tb.Button(
            auto_input,
            text="追加 / 上書き",
            command=self._add_auto_map,
            bootstyle=SECONDARY,
        ).pack(side=tk.LEFT)
        tb.Button(
            auto_input,
            text="選択を削除",
            command=self._delete_selected_auto_map,
            bootstyle="outline-secondary",
        ).pack(side=tk.LEFT, padx=6)

        self.auto_map_tree = ttk.Treeview(
            auto_map_frame,
            columns=("measure_no", "data_index"),
            show="headings",
            height=6,
            style="Data.Treeview",
        )
        self.auto_map_tree.heading("measure_no", text="測定 No")
        self.auto_map_tree.heading("data_index", text="データの順番")
        self.auto_map_tree.column("measure_no", width=200, anchor=tk.W)
        self.auto_map_tree.column("data_index", width=120, anchor=tk.CENTER)
        self.auto_map_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        auto_scrollbar = ttk.Scrollbar(
            auto_map_frame,
            orient="vertical",
            command=self.auto_map_tree.yview,
        )
        self.auto_map_tree.configure(yscrollcommand=auto_scrollbar.set)
        auto_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        action_bar = ttk.LabelFrame(
            main,
            text="実行",
            style="AppCard.TLabelframe",
            padding=(14, 14, 14, 14),
        )
        action_bar.pack(fill=tk.X)
        lower = ttk.Frame(action_bar, style="Surface.TFrame")
        lower.pack(fill=tk.X)
        ttk.Label(
            lower,
            text="元の Excel・保存先のファイルは、必ず保存して閉じてから実行してください。",
            style="CardNote.TLabel",
            wraplength=480,
            justify=tk.LEFT,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        tb.Button(
            lower,
            text="　この内容で Excel を保存・生成　",
            command=self._run_build,
            bootstyle=INFO,
        ).pack(side=tk.RIGHT)

        if not self.tools_tree.get_children():
            self._insert_tool("前挽き(サンプル)", "1, 5, 10")

    def _is_in_main_content(self, widget):
        current = widget
        while current is not None:
            if current is self.main_canvas:
                return True
            parent_name = current.winfo_parent()
            if not parent_name:
                break
            try:
                current = current.nametowidget(parent_name)
            except Exception:
                break
        return False

    def _on_main_mousewheel(self, event):
        if not self._is_in_main_content(event.widget):
            return
        if getattr(event, "delta", 0):
            self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif getattr(event, "num", None) == 4:
            self.main_canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.main_canvas.yview_scroll(1, "units")

    def _bind_basic_setting_sync(self):
        self.vars["not_required_row"].trace_add("write", self._sync_not_required_row)
        self.vars["measure_row_min"].trace_add("write", self._sync_measure_row_min)
        self.vars["measure_row_max"].trace_add("write", self._sync_measure_row_max)
        self.vars["measure_row_step"].trace_add("write", self._sync_measure_row_step)

    def _apply_locked_basic_settings(self):
        for key, value in LOCKED_BASIC_SETTINGS.items():
            if self.vars[key].get() != value:
                self.vars[key].set(value)

    def _sync_not_required_row(self, *args):
        try:
            not_required_row = int(self.vars["not_required_row"].get())
        except (tk.TclError, ValueError):
            return
        min_row = self.vars["measure_row_min"].get()
        desired_max, desired_tool_start = _derive_layout_rows(not_required_row, min_row)
        if self.vars["measure_row_max"].get() != desired_max:
            self.vars["measure_row_max"].set(desired_max)
        if self.vars["tool_start_row"].get() != desired_tool_start:
            self.vars["tool_start_row"].set(desired_tool_start)

    def _sync_measure_row_min(self, *args):
        try:
            min_row = self.vars["measure_row_min"].get()
        except tk.TclError:
            return
        self.vars["summary_row_min"].set(min_row)
        if self.vars["measure_row_max"].get() < min_row:
            self.vars["measure_row_max"].set(min_row)

    def _sync_measure_row_max(self, *args):
        try:
            max_row = self.vars["measure_row_max"].get()
        except tk.TclError:
            return
        self.vars["summary_row_max"].set(max_row)

    def _sync_measure_row_step(self, *args):
        try:
            step = self.vars["measure_row_step"].get()
        except tk.TclError:
            return
        if step < 1:
            self.vars["measure_row_step"].set(1)
            return
        self.vars["summary_row_step"].set(step)
        if self.vars["tool_row_step"].get() != step:
            self.vars["tool_row_step"].set(step)

    def _load_preview(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="元の検査シート（xlsx）を選択",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        self.selected_xlsx.set(path)

        loading = LoadingDialog(self, "読み込み中...", "Excelファイルを読み込んでいます...")
        result = {"success": False, "error": None}

        def load_task():
            try:
                self._render_preview_internal()
                result["success"] = True
            except Exception as e:
                result["error"] = e

        thread = threading.Thread(target=load_task, daemon=True)
        thread.start()

        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
                return
            loading.close()
            if result["error"]:
                messagebox.showerror("読み込み失敗", str(result["error"]), parent=self)

        self.after(100, check_completion)

    def _render_preview(self):
        path = self.selected_xlsx.get().strip()
        if not path:
            messagebox.showinfo("読み込み待ち", "先にExcelファイルを選択してください。", parent=self)
            return

        loading = LoadingDialog(self, "更新中...", "プレビューを更新しています...")
        result = {"success": False, "error": None}

        def render_task():
            try:
                self._render_preview_internal()
                result["success"] = True
            except Exception as e:
                result["error"] = e

        thread = threading.Thread(target=render_task, daemon=True)
        thread.start()

        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
                return
            loading.close()
            if result["error"]:
                messagebox.showerror("更新失敗", str(result["error"]), parent=self)

        self.after(100, check_completion)

    def _render_preview_internal(self):
        path = self.selected_xlsx.get().strip()
        if not path:
            return

        sheet_name = self.vars["sheet_name"].get().strip() or "工程内検査シート"
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            raise Exception(f"Excelを開けませんでした。\n{e}")

        if sheet_name not in wb.sheetnames:
            raise Exception(
                f"シート「{sheet_name}」が見つかりません。\nシート名を確認して再度プレビューしてください。"
            )

        ws = wb[sheet_name]
        preview_col_indices = [column_index_from_string(column) for column in self.preview_columns]

        if not ws.max_row or not ws.max_column:
            raise Exception("表示できるデータがありません。")

        col_widths = {"A": 80, "B": 120, "G": 140, "K": 140}

        def update_ui():
            self.preview_tree.configure(columns=self.preview_columns)
            for col in self.preview_columns:
                self.preview_tree.heading(col, text=col, anchor="center")
                self.preview_tree.column(
                    col,
                    width=col_widths.get(col, 120),
                    minwidth=60,
                    anchor="center",
                    stretch=False,
                )

            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)

            header_row = 10
            group_size = 3
            max_row_value = ws.max_row or 0

            if max_row_value < header_row:
                raise Exception(f"{header_row}行目以降に表示できるデータがありません。")

            header_vals = []
            for col_idx in preview_col_indices:
                value = ws.cell(header_row, col_idx).value
                header_vals.append("" if value is None else str(value))
            self.preview_tree.insert("", "end", values=header_vals)

            data_start_row = header_row + 1
            row_count = 0
            for row_index in range(data_start_row, max_row_value + 1, group_size):
                a_val = ws.cell(row_index, preview_col_indices[0]).value
                if a_val is None or (isinstance(a_val, str) and not a_val.strip()):
                    break

                row_vals = []
                for col_idx in preview_col_indices:
                    value = ws.cell(row_index, col_idx).value
                    row_vals.append("" if value is None else str(value))

                self.preview_tree.insert("", "end", values=row_vals)
                row_count += 1

            if row_count == 0:
                for item in self.preview_tree.get_children():
                    self.preview_tree.delete(item)
                raise Exception("11行目以降の先頭行(A列)が空のため表示できるデータがありません。")

            self.preview_title.set(
                f"{sheet_name} プレビュー（{row_count}行: 10行目ヘッダー＋11行目以降3行刻み先頭行）"
            )

        self.after(0, update_ui)

    def _insert_tool(self, tool_name: str, nos_text: str):
        self.tools_tree.insert("", "end", values=(tool_name, nos_text))

    def _tool_dialog(self, title, init_tool="", init_nos=""):
        win = tb.Toplevel(self)
        win.title(title)
        win.transient(self)
        win.grab_set()
        try:
            win.configure(bg=self["bg"])
        except tk.TclError:
            pass

        tool_var = tk.StringVar(value=init_tool)
        nos_var = tk.StringVar(value=init_nos)

        frm = ttk.Frame(win, style="Surface.TFrame", padding=22)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text="工具名", style="CardTitle.TLabel").grid(row=0, column=0, sticky=tk.W, pady=4)
        tool_entry = ttk.Entry(frm, textvariable=tool_var, width=36)
        tool_entry.grid(row=0, column=1, sticky=tk.EW, pady=4, padx=(8, 0))

        ttk.Label(frm, text="測定 No", style="CardTitle.TLabel").grid(row=1, column=0, sticky=tk.NW, pady=10)
        nos_box = ttk.Frame(frm, style="Surface.TFrame")
        nos_box.grid(row=1, column=1, sticky=tk.EW, pady=10, padx=(8, 0))
        ttk.Entry(nos_box, textvariable=nos_var, width=36).pack(anchor=tk.W)
        ttk.Label(
            nos_box,
            text="半角の整数をカンマで区切ります。例: 1, 5, 10",
            style="CardNote.TLabel",
            wraplength=320,
        ).pack(anchor=tk.W, pady=(4, 0))
        frm.columnconfigure(1, weight=1)

        result = {"ok": False}

        def on_ok():
            tool = tool_var.get().strip()
            if not tool:
                messagebox.showwarning("入力不足", "工具名を入力してください。", parent=win)
                return
            try:
                _parse_int_list(nos_var.get())
            except Exception:
                messagebox.showwarning(
                    "入力エラー",
                    "測定 No は、整数をカンマ区切りで入力してください。",
                    parent=win,
                )
                return
            result["ok"] = True
            result["tool"] = tool
            result["nos"] = nos_var.get().strip()
            win.destroy()

        def on_cancel():
            win.destroy()

        bfrm = ttk.Frame(frm, style="Surface.TFrame")
        bfrm.grid(row=2, column=0, columnspan=2, sticky=tk.E, pady=(20, 0))
        tb.Button(bfrm, text="キャンセル", command=on_cancel, bootstyle="outline-secondary").pack(
            side=tk.RIGHT, padx=(6, 0)
        )
        tb.Button(bfrm, text="登録", command=on_ok, bootstyle=INFO).pack(side=tk.RIGHT)

        tool_entry.focus_set()
        win.update_idletasks()
        _tw, _th = 580, 300
        win.minsize(_tw, _th)
        place_toplevel_center(win, _tw, _th)
        self.wait_window(win)
        return result

    def _add_tool_dialog(self):
        result = self._tool_dialog("工具追加")
        if result.get("ok"):
            self._insert_tool(result["tool"], result["nos"])

    def _edit_selected_tool(self):
        selected = self.tools_tree.selection()
        if not selected:
            messagebox.showinfo("選択なし", "編集する行を選択してください。", parent=self)
            return
        item = selected[0]
        tool, nos = self.tools_tree.item(item, "values")
        result = self._tool_dialog("工具編集", init_tool=tool, init_nos=nos)
        if result.get("ok"):
            self.tools_tree.item(item, values=(result["tool"], result["nos"]))

    def _delete_selected_tool(self):
        selected = self.tools_tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("削除確認", "選択した工具を削除しますか？", parent=self):
            return
        for item in selected:
            self.tools_tree.delete(item)

    def _add_auto_map(self):
        measure_no_raw = self.auto_map_measure_no_var.get().strip()
        data_index_raw = self.auto_map_data_index_var.get().strip()

        key = _normalize_measure_no_key(measure_no_raw)
        data_index = _try_extract_int(data_index_raw)

        if key == "":
            messagebox.showwarning("入力エラー", "測定Noを入力してください。", parent=self)
            return
        if data_index is None or data_index < 1 or data_index > AUTO_DATA_MAX_ITEMS:
            messagebox.showwarning(
                "入力エラー",
                f"データ順番は1〜{AUTO_DATA_MAX_ITEMS}の整数で入力してください。",
                parent=self,
            )
            return

        key_str = str(key)
        existing = None
        for item in self.auto_map_tree.get_children():
            values = self.auto_map_tree.item(item, "values")
            if str(values[0]) == key_str:
                existing = item
                break

        if existing is not None:
            self.auto_map_tree.item(existing, values=(key_str, str(data_index)))
        else:
            self.auto_map_tree.insert("", "end", values=(key_str, str(data_index)))

        self.auto_map_measure_no_var.set("")
        self.auto_map_data_index_var.set("")

    def _delete_selected_auto_map(self):
        selected = self.auto_map_tree.selection()
        if not selected:
            return
        for item in selected:
            self.auto_map_tree.delete(item)

    def _collect_not_required_nos(self) -> list[int]:
        nums: list[int] = []
        for item in self.not_required_nos_tree.get_children():
            try:
                n = int(self.not_required_nos_tree.item(item, "values")[0])
            except (TypeError, ValueError, IndexError, tk.TclError):
                continue
            nums.append(n)
        return sorted(set(nums))

    def _add_not_required_no(self):
        raw = self.not_required_no_input_var.get().strip()
        n = _try_extract_int(raw)
        if n is None:
            messagebox.showwarning("入力エラー", "測定 No を整数で入力してください。", parent=self)
            return
        for item in self.not_required_nos_tree.get_children():
            try:
                v = int(self.not_required_nos_tree.item(item, "values")[0])
            except (TypeError, ValueError, IndexError, tk.TclError):
                continue
            if v == n:
                messagebox.showinfo("重複", "この No はすでに登録されています。", parent=self)
                return
        self.not_required_nos_tree.insert("", "end", values=(str(n),))
        self._sort_not_required_nos_tree()
        self.not_required_no_input_var.set("")

    def _sort_not_required_nos_tree(self):
        values: list[int] = []
        for item in self.not_required_nos_tree.get_children():
            try:
                values.append(int(self.not_required_nos_tree.item(item, "values")[0]))
            except (TypeError, ValueError, IndexError, tk.TclError):
                continue
        for item in self.not_required_nos_tree.get_children():
            self.not_required_nos_tree.delete(item)
        for n in sorted(set(values)):
            self.not_required_nos_tree.insert("", "end", values=(str(n),))

    def _delete_selected_not_required_no(self):
        selected = self.not_required_nos_tree.selection()
        if not selected:
            messagebox.showinfo("選択なし", "削除する行を選んでください。", parent=self)
            return
        for item in selected:
            self.not_required_nos_tree.delete(item)

    def _gather_cfg(self):
        try:
            self._apply_locked_basic_settings()
            tools = []
            tool_to_measure_nos = {}
            measure_no_to_data_index = {}

            for item in self.tools_tree.get_children():
                tool, nos_text = self.tools_tree.item(item, "values")
                tools.append(tool)
                tool_to_measure_nos[tool] = _parse_int_list(nos_text)

            for item in self.auto_map_tree.get_children():
                measure_no_text, data_index_text = self.auto_map_tree.item(item, "values")
                key = _normalize_measure_no_key(measure_no_text)
                data_index = _try_extract_int(data_index_text)
                if key == "" or data_index is None:
                    continue
                measure_no_to_data_index[key] = data_index

            if not tools:
                raise ValueError("工具が1件もありません。")

            not_required_row_text = self.vars["not_required_row"].get().strip()
            not_required_row = _try_extract_int(not_required_row_text)
            if not_required_row is None:
                raise ValueError("測定不要書き込み設定の行は整数で入力してください。")

            measure_row_min = LOCKED_BASIC_SETTINGS["measure_row_min"]
            measure_row_max, tool_start_row = _derive_layout_rows(not_required_row, measure_row_min)
            if measure_row_max < measure_row_min:
                raise ValueError(
                    f"測定不要書き込み設定の行は {measure_row_min + 1} 以上で入力してください。"
                )
            if tool_start_row < 1:
                raise ValueError("自動計算後の工具開始行が1未満になります。入力値を見直してください。")
            auto_data_start_row = _derive_auto_data_start_row(not_required_row, len(tools))

            cfg = {
                "sheet_name": self.vars["sheet_name"].get().strip(),
                "measure_no_col": LOCKED_BASIC_SETTINGS["measure_no_col"],
                "measure_row_min": measure_row_min,
                "measure_row_max": measure_row_max,
                "measure_row_step": LOCKED_BASIC_SETTINGS["measure_row_step"],
                "summary_row_min": LOCKED_BASIC_SETTINGS["summary_row_min"],
                "summary_row_max": measure_row_max,
                "summary_row_step": LOCKED_BASIC_SETTINGS["summary_row_step"],
                "formula_arg_sep": LOCKED_BASIC_SETTINGS["formula_arg_sep"],
                "tool_start_row": tool_start_row,
                "not_required_row": not_required_row,
                "tool_name_col": LOCKED_BASIC_SETTINGS["tool_name_col"],
                "tool_row_step": LOCKED_BASIC_SETTINGS["tool_row_step"],
                "auto_data_start_row": auto_data_start_row,
                "measure_no_to_data_index": measure_no_to_data_index,
                "tools": tools,
                "tool_to_measure_nos": tool_to_measure_nos,
            }
            if not cfg["sheet_name"]:
                raise ValueError("シート名が空です。")
            return cfg
        except Exception as e:
            raise ValueError(f"設定の取得に失敗: {e}")

    def _run_build(self):
        try:
            cfg = self._gather_cfg()
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return

        xlsx = self.selected_xlsx.get().strip()
        if not xlsx:
            messagebox.showinfo(
                "ファイル未選択",
                "先に「参照…」で元の Excel を選んでください。",
                parent=self,
            )
            self._load_preview()
            xlsx = self.selected_xlsx.get().strip()
            if not xlsx:
                return

        out_path = pick_save_path(
            "出力先（生成したxlsx）を保存",
            ".xlsx",
            [("Excel", "*.xlsx")],
            parent=self,
        )
        if not out_path:
            return

        loading = LoadingDialog(self, "生成中...", "Excelファイルを生成しています...")
        result = {"success": False, "error": None, "saved_path": None}

        def build_task():
            try:
                saved_path = build_request_formulas(xlsx, out_path, cfg, parent=self)
                result["saved_path"] = saved_path

                target_nos = self._collect_not_required_nos()
                if target_nos:
                    try:
                        saved_path = write_measurement_not_required(
                            saved_path,
                            out_path,
                            cfg,
                            target_nos=target_nos,
                            parent=self,
                        )
                        result["saved_path"] = saved_path
                    except Exception as e:
                        result["error"] = f"測定不要書き込みでエラーが発生しました:\n{e}"
                        result["success"] = True
                        return

                result["success"] = True
            except Exception as e:
                result["error"] = str(e)

        thread = threading.Thread(target=build_task, daemon=True)
        thread.start()

        def check_completion():
            if thread.is_alive():
                self.after(100, check_completion)
                return
            loading.close()
            if result["success"]:
                if result["error"]:
                    messagebox.showwarning(
                        "警告",
                        f"Excel生成は完了しましたが、{result['error']}",
                        parent=self,
                    )
                else:
                    messagebox.showinfo("完了", f"生成しました:\n{result['saved_path']}", parent=self)
                return
            messagebox.showerror("失敗", result["error"], parent=self)

        self.after(100, check_completion)

    def _run_write_not_required(self):
        try:
            cfg = self._gather_cfg()
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return

        target_nos = self._collect_not_required_nos()
        if not target_nos:
            messagebox.showwarning(
                "入力不足",
                "L～SR で「-」にする測定 No を1件以上登録してください。",
                parent=self,
            )
            return

        xlsx = self.selected_xlsx.get().strip()
        if not xlsx:
            messagebox.showinfo(
                "ファイル未選択",
                "先に「参照…」で元の Excel を選んでください。",
                parent=self,
            )
            self._load_preview()
            xlsx = self.selected_xlsx.get().strip()
            if not xlsx:
                return
        out_path = pick_save_path(
            "出力先（生成したxlsx）を保存",
            ".xlsx",
            [("Excel", "*.xlsx")],
            parent=self,
        )
        if not out_path:
            return

        try:
            saved_path = write_measurement_not_required(
                xlsx,
                out_path,
                cfg,
                target_nos=target_nos,
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("失敗", str(e), parent=self)
            return
        messagebox.showinfo("完了", f"書き込み完了:\n{saved_path}", parent=self)

    def _show_help(self):
        open_help_window(self)


def main():
    app = ConfigEditor()
    app.mainloop()
