import copy
import os
import posixpath
import re
import tempfile
import time
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from tkinter import messagebox


# 測定不要行デフォルト 122。配下の自動データ開始行フォールバックは 1 工具想定: 122 + 3 + 6
AUTO_DATA_START_ROW_DEFAULT = 131
NOT_REQUIRED_ROW_DEFAULT = 122
AUTO_DATA_MAX_ITEMS = 100
REQUEST_HEADER_ROW = 10
SUMMARY_FORMULA_COL_START = "L"
SUMMARY_FORMULA_COL_END = "SN"
REQUEST_OUTPUT_COL_START = "L"
REQUEST_OUTPUT_COL_END = "SR"
# 1〜3 行目の SUMPRODUCT: 行ごとに開始/終了が 1 行ずつずれる（L11:L119 / L12:L120 / L13:L121）
SUMMARY_FORMULA_BASE_START_ROW = 11
SUMMARY_FORMULA_BASE_END_ROW = 119
SUMMARY_FORMULA_MOD_DIVISOR = 3
LOCKED_BASIC_SETTINGS = {
    "measure_no_col": "A",
    "measure_row_min": 11,
    "measure_row_step": 3,
    "summary_row_min": 11,
    "summary_row_step": 3,
    "formula_arg_sep": ",",
    "tool_name_col": "E",
    "tool_row_step": 3,
}
FORCE_EXCEL_RECALC_ENV = "FLAG_AUTO_GENERATOR_FORCE_EXCEL_RECALC"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
XR2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
XR3_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
X15_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
X15AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac"
XR6_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6"
XR10_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10"
XCALCF_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures"
NS = {
    "main": MAIN_NS,
    "rel": REL_NS,
    "pkg": PKG_REL_NS,
    "ct": CONTENT_TYPES_NS,
}

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)
ET.register_namespace("mc", MC_NS)
ET.register_namespace("x14ac", X14AC_NS)
ET.register_namespace("xr", XR_NS)
ET.register_namespace("xr2", XR2_NS)
ET.register_namespace("xr3", XR3_NS)
ET.register_namespace("x15", X15_NS)
ET.register_namespace("x15ac", X15AC_NS)
ET.register_namespace("xr6", XR6_NS)
ET.register_namespace("xr10", XR10_NS)
ET.register_namespace("xcalcf", XCALCF_NS)


def _derive_layout_rows(not_required_row: int, measure_row_min: int) -> tuple[int, int]:
    measure_row_max = max(not_required_row - 1, measure_row_min)
    tool_start_row = max(not_required_row + 3, 1)
    return measure_row_max, tool_start_row


def _derive_auto_data_start_row(not_required_row: int, tool_count: int) -> int:
    return not_required_row + (tool_count * LOCKED_BASIC_SETTINGS["tool_row_step"]) + 6


def _get_writable_cell(ws, row: int, col: int):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col)

    return cell


def _save_workbook_to_temp_file(wb, out_path_hint: str) -> str:
    _, ext = os.path.splitext(out_path_hint)
    ext = ext or ".xlsx"
    fd, tmp_path = tempfile.mkstemp(suffix=ext)
    os.close(fd)
    wb.save(tmp_path)
    return tmp_path


def _replace_file_atomic(temp_path: str, out_path: str, parent=None) -> str:
    out_path = os.path.abspath(out_path)
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    while True:
        try:
            os.replace(temp_path, out_path)
            return out_path
        except PermissionError as e:
            if parent is None:
                raise
            retry = messagebox.askretrycancel(
                "保存に失敗",
                "出力先ファイルに書き込めません。\n"
                "Excelで出力先ファイルを開いている場合は閉じてから「再試行」を押してください。\n\n"
                f"出力先:\n{out_path}\n\n"
                f"詳細:\n{e}",
                parent=parent,
            )
            if not retry:
                raise


def _mark_workbook_for_full_recalc(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _safe_call(func, *args, default=None, **kwargs):
    try:
        return func(*args, **kwargs)
    except Exception:
        return default


def _close_workbook_quietly(wb):
    if wb is None:
        return
    _safe_call(wb.close)


def _normalize_package_path(target: str) -> str:
    normalized = (target or "").replace("\\", "/")
    if normalized.startswith("/"):
        normalized = normalized[1:]
    return posixpath.normpath(normalized)


def _worksheet_path_by_name(xlsx_path: str, sheet_name: str) -> str:
    with zipfile.ZipFile(xlsx_path, "r") as workbook_zip:
        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        workbook_rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))

    rel_id_to_target = {}
    for rel in workbook_rels_root.findall("pkg:Relationship", NS):
        rel_id = rel.attrib.get("Id")
        target = _normalize_package_path(rel.attrib.get("Target", ""))
        if rel_id:
            rel_id_to_target[rel_id] = target

    for sheet in workbook_root.findall("main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
        if not rel_id:
            break
        target = rel_id_to_target.get(rel_id, "")
        if target.startswith("xl/"):
            return target
        if target.startswith("worksheets/"):
            return f"xl/{target}"
        return f"xl/{target}"

    raise ValueError(f"シート '{sheet_name}' の XML パスを特定できませんでした。")


def _cell_ref_sort_key(cell_ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", cell_ref.upper())
    if not match:
        return (10**9, 10**9)
    col_letters, row_text = match.groups()
    return (int(row_text), column_index_from_string(col_letters))


def _find_or_create_row(sheet_data, row_index: int):
    existing_rows = sheet_data.findall("main:row", NS)
    for row in existing_rows:
        if int(row.attrib.get("r", "0")) == row_index:
            return row

    new_row = ET.Element(f"{{{MAIN_NS}}}row", {"r": str(row_index)})
    inserted = False
    for pos, row in enumerate(existing_rows):
        if int(row.attrib.get("r", "0")) > row_index:
            sheet_data.insert(pos, new_row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(new_row)
    return new_row


def _set_row_cell(row_elem, cell_ref: str, new_cell):
    new_cell_copy = copy.deepcopy(new_cell)
    existing_cells = row_elem.findall("main:c", NS)
    for pos, cell in enumerate(existing_cells):
        current_ref = cell.attrib.get("r", "")
        if current_ref == cell_ref:
            row_elem.remove(cell)
            row_elem.insert(pos, new_cell_copy)
            return
        if _cell_ref_sort_key(current_ref) > _cell_ref_sort_key(cell_ref):
            row_elem.insert(pos, new_cell_copy)
            return
    row_elem.append(new_cell_copy)


def _merge_sheet_cells(source_xml: bytes, modified_xml: bytes, changed_refs: set[str]) -> bytes:
    if not changed_refs:
        return source_xml

    source_root = ET.fromstring(source_xml)
    modified_root = ET.fromstring(modified_xml)
    source_sheet_data = source_root.find("main:sheetData", NS)
    modified_sheet_data = modified_root.find("main:sheetData", NS)
    if source_sheet_data is None or modified_sheet_data is None:
        raise ValueError("sheetData の解析に失敗しました。")

    modified_cell_map = {}
    for row in modified_sheet_data.findall("main:row", NS):
        for cell in row.findall("main:c", NS):
            cell_ref = cell.attrib.get("r")
            if cell_ref:
                modified_cell_map[cell_ref] = cell

    for cell_ref in sorted(changed_refs, key=_cell_ref_sort_key):
        modified_cell = modified_cell_map.get(cell_ref)
        if modified_cell is None:
            continue
        row_index = _cell_ref_sort_key(cell_ref)[0]
        target_row = _find_or_create_row(source_sheet_data, row_index)
        _set_row_cell(target_row, cell_ref, modified_cell)

    merged_xml = ET.tostring(source_root, encoding="utf-8", xml_declaration=True)
    return _restore_root_namespace_declarations(merged_xml, source_xml)


def _mark_workbook_xml_for_full_recalc(workbook_xml: bytes) -> bytes:
    root = ET.fromstring(workbook_xml)
    calc_pr = root.find("main:calcPr", NS)
    if calc_pr is None:
        calc_pr = ET.SubElement(root, f"{{{MAIN_NS}}}calcPr")
    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    serialized_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return _restore_root_namespace_declarations(serialized_xml, workbook_xml)


def _restore_root_namespace_declarations(serialized_xml: bytes, original_xml: bytes) -> bytes:
    serialized_text = serialized_xml.decode("utf-8")
    original_text = original_xml.decode("utf-8", errors="ignore")

    serialized_match = re.search(r"<([A-Za-z0-9_:.-]+)([^>]*)>", serialized_text)
    original_match = re.search(r"<([A-Za-z0-9_:.-]+)([^>]*)>", original_text)
    if not serialized_match or not original_match:
        return serialized_xml

    serialized_root_tag = serialized_match.group(1)
    serialized_root_attrs = serialized_match.group(2)
    original_root_attrs = original_match.group(2)

    namespace_decls = re.findall(r'\s(xmlns(?::[A-Za-z0-9_.-]+)?)="([^"]+)"', original_root_attrs)
    missing_decls = []
    for attr_name, uri in namespace_decls:
        decl_pattern = rf'\s{re.escape(attr_name)}="{re.escape(uri)}"'
        if re.search(decl_pattern, serialized_root_attrs):
            continue
        missing_decls.append(f' {attr_name}="{uri}"')

    if not missing_decls:
        return serialized_xml

    replacement = f"<{serialized_root_tag}{serialized_root_attrs}{''.join(missing_decls)}>"
    updated_text = (
        serialized_text[: serialized_match.start()]
        + replacement
        + serialized_text[serialized_match.end() :]
    )
    return updated_text.encode("utf-8")


def _remove_calc_chain_parts(package_files: dict[str, bytes]):
    package_files.pop("xl/calcChain.xml", None)

    content_types_xml = package_files.get("[Content_Types].xml")
    if content_types_xml is not None:
        content_types_root = ET.fromstring(content_types_xml)
        for override in list(content_types_root.findall("ct:Override", NS)):
            if override.attrib.get("PartName") == "/xl/calcChain.xml":
                content_types_root.remove(override)
        package_files["[Content_Types].xml"] = ET.tostring(
            content_types_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    workbook_rels_xml = package_files.get("xl/_rels/workbook.xml.rels")
    if workbook_rels_xml is not None:
        workbook_rels_root = ET.fromstring(workbook_rels_xml)
        for rel in list(workbook_rels_root.findall("pkg:Relationship", NS)):
            if rel.attrib.get("Type", "").endswith("/calcChain"):
                workbook_rels_root.remove(rel)
        package_files["xl/_rels/workbook.xml.rels"] = ET.tostring(
            workbook_rels_root,
            encoding="utf-8",
            xml_declaration=True,
        )


def _save_preserving_package_parts(
    source_xlsx_path: str,
    modified_xlsx_path: str,
    out_path: str,
    sheet_name: str,
    changed_refs: set[str],
    *,
    parent=None,
) -> str:
    source_sheet_path = _worksheet_path_by_name(source_xlsx_path, sheet_name)
    modified_sheet_path = _worksheet_path_by_name(modified_xlsx_path, sheet_name)
    out_path = os.path.abspath(out_path)
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    base, ext = os.path.splitext(out_path)
    ext = ext or ".xlsx"
    temp_out_path = f"{base}.tmp{ext}"

    with zipfile.ZipFile(source_xlsx_path, "r") as source_zip:
        package_files = {
            name: source_zip.read(name)
            for name in source_zip.namelist()
        }

    with zipfile.ZipFile(modified_xlsx_path, "r") as modified_zip:
        merged_sheet_xml = _merge_sheet_cells(
            package_files[source_sheet_path],
            modified_zip.read(modified_sheet_path),
            changed_refs,
        )
        package_files["xl/styles.xml"] = modified_zip.read("xl/styles.xml")

    package_files[source_sheet_path] = merged_sheet_xml
    package_files["xl/workbook.xml"] = _mark_workbook_xml_for_full_recalc(
        package_files["xl/workbook.xml"]
    )
    _remove_calc_chain_parts(package_files)

    try:
        with zipfile.ZipFile(temp_out_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
            for name, data in package_files.items():
                out_zip.writestr(name, data)
        return _replace_file_atomic(temp_out_path, out_path, parent=parent)
    finally:
        if os.path.exists(temp_out_path):
            _safe_call(os.remove, temp_out_path)


def _finalize_modified_workbook(
    wb,
    wb_values,
    *,
    source_xlsx_path: str,
    out_path: str,
    sheet_name: str,
    changed_refs: set[str],
    parent=None,
) -> str:
    _mark_workbook_for_full_recalc(wb)
    temp_saved_path = _save_workbook_to_temp_file(wb, out_path)
    _close_workbook_quietly(wb_values)
    _close_workbook_quietly(wb)

    try:
        return _save_preserving_package_parts(
            source_xlsx_path,
            temp_saved_path,
            out_path,
            sheet_name,
            changed_refs,
            parent=parent,
        )
    finally:
        _safe_call(os.remove, temp_saved_path)


def _force_excel_recalc_and_save(xlsx_path: str):
    if os.environ.get(FORCE_EXCEL_RECALC_ENV, "").strip() != "1":
        print(
            "[info] Excel強制再計算は既定でスキップします。"
            f"必要な場合のみ環境変数 {FORCE_EXCEL_RECALC_ENV}=1 で有効化してください。"
        )
        return False

    try:
        import win32com.client  # type: ignore
    except Exception as e:
        print(f"[warn] Excel強制再計算をスキップしました（pywin32未導入）: {e}")
        return False

    abs_path = os.path.abspath(xlsx_path)
    last_error = None

    for attempt in range(3):
        excel_app = None
        excel_wb = None
        try:
            print(f"[info] Excel強制再計算を開始します ({attempt + 1}/3)")
            excel_app = win32com.client.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            _safe_call(setattr, excel_app, "AskToUpdateLinks", False)

            excel_wb = excel_app.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=False)

            _safe_call(setattr, excel_wb, "ForceFullCalculation", True)

            worksheets = _safe_call(lambda: list(excel_wb.Worksheets), default=[])
            for sheet in worksheets:
                _safe_call(setattr, sheet, "EnableCalculation", True)
                _safe_call(sheet.UsedRange.Calculate)

            _safe_call(excel_wb.RefreshAll)
            _safe_call(excel_app.CalculateFull)
            excel_app.CalculateFullRebuild()
            save_as_kwargs = {
                "Filename": abs_path,
                "FileFormat": 51,
                "ConflictResolution": 2,
                "Local": True,
            }
            if _safe_call(excel_wb.SaveAs, **save_as_kwargs) is None:
                excel_wb.Save()
            print("[info] Excel強制再計算が完了しました")
            return True
        except Exception as e:
            last_error = e
            time.sleep(0.4 * (attempt + 1))
        finally:
            if excel_wb is not None:
                _safe_call(excel_wb.Close, SaveChanges=True)
            if excel_app is not None:
                _safe_call(excel_app.Quit)

    print(f"[info] Excel強制再計算をスキップしました（出力ファイルは作成済み）: {last_error}")
    return False


def _try_extract_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        try:
            return int(value)
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(s)
        except Exception:
            m = re.search(r"\d+", s)
            if not m:
                return None
            try:
                return int(m.group(0))
            except Exception:
                return None
    return None


def _resolve_measure_no(value, row: int, measure_row_min: int, measure_row_step: int):
    if isinstance(value, str) and value.startswith("="):
        if measure_row_step <= 0:
            return None

        row_offset = row - measure_row_min
        if row_offset < 0:
            return None
        if row_offset % measure_row_step != 0:
            return None

        return (row_offset // measure_row_step) + 1

    return _try_extract_int(value)


def _normalize_measure_no_key(value):
    n = _try_extract_int(value)
    if n is not None:
        return n
    if value is None:
        return ""
    return str(value).strip()


def _normalize_measure_to_index_map(raw_map: dict):
    normalized = {}
    if not isinstance(raw_map, dict):
        return normalized
    for measure_no, data_index in raw_map.items():
        key = _normalize_measure_no_key(measure_no)
        if key == "":
            continue
        index_value = _try_extract_int(data_index)
        if index_value is None or index_value < 1 or index_value > AUTO_DATA_MAX_ITEMS:
            continue
        normalized[key] = index_value
    return normalized


def _build_auto_data_formula(col_letter: str, data_start_row: int, data_index: int, sep: str):
    source_row = data_start_row + data_index - 1
    return (
        f'IFERROR(IF({col_letter}{source_row}=""{sep}""{sep}{col_letter}{source_row}){sep}"")'
    )


def _is_empty_cell_value(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _to_formula_expression(value):
    if _is_empty_cell_value(value):
        return '""'
    if isinstance(value, str) and value.startswith("="):
        return value[1:]
    if isinstance(value, str):
        escaped = value.replace('"', '""')
        return f'"{escaped}"'
    return str(value)


def _build_not_required_overlay_formula(
    current_value,
    trigger_cell_ref: str,
    sep: str,
):
    base_expression = _to_formula_expression(current_value)
    return (
        f'=IF(IFERROR(LEN(TRIM(({base_expression})&"")){sep}0)>0{sep}'
        f'({base_expression}){sep}'
        f'IF(IFERROR(LEN(TRIM({trigger_cell_ref}&"")){sep}0)>0{sep}"-"{sep}""))'
    )


def _can_overwrite_with_formula(value):
    if _is_empty_cell_value(value):
        return True
    if isinstance(value, str) and value.startswith("="):
        return True
    return False


def _normalize_single_cell_array_formulas_in_column(
    ws,
    col_letter: str,
    *,
    row_start: int = 1,
    row_end: int | None = None,
):
    col_idx = column_index_from_string(col_letter)
    max_row = ws.max_row or row_start
    end_row = max_row if row_end is None else min(row_end, max_row)
    rewritten_refs = set()

    start_row = max(row_start, 1)
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row, col_idx)
        value = cell.value
        if not isinstance(value, ArrayFormula):
            continue

        formula_ref = str(value.ref or "").replace("$", "")
        cell_ref = cell.coordinate.replace("$", "")
        if formula_ref != cell_ref:
            continue

        formula_text = str(value.text or "").strip()
        if not formula_text:
            continue

        cell.value = f"={formula_text.lstrip('=')}"
        rewritten_refs.add(cell.coordinate)

    return rewritten_refs


def _build_summary_formula(col_letter: str, row_offset: int) -> str:
    start_row = SUMMARY_FORMULA_BASE_START_ROW + row_offset
    end_row = SUMMARY_FORMULA_BASE_END_ROW + row_offset
    m = SUMMARY_FORMULA_MOD_DIVISOR
    return (
        f"=SUMPRODUCT(--({col_letter}{start_row}:{col_letter}{end_row}<>\"\"),"
        f"--(MOD(ROW({col_letter}{start_row}:{col_letter}{end_row})-ROW({col_letter}{start_row}),{m})=0))"
    )


def _build_stepped_non_empty_count_formula(
    col_letter: str,
    row_min: int,
    row_max: int,
    row_step: int,
    sep: str,
) -> str:
    target_range = f"{col_letter}{row_min}:{col_letter}{row_max}"
    start_ref = f"{col_letter}{row_min}"
    return (
        f'SUMPRODUCT(--(LEN(TRIM({target_range}&""))>0){sep}'
        f'--(MOD(ROW({target_range})-ROW({start_ref}){sep}{row_step})=0))'
    )


def _build_request_formula(conditions: str, fallback_expression: str = '""') -> str:
    return f'=IF(OR({conditions}),"依頼",{fallback_expression})'


def _build_measure_row_formula(
    request_conditions: str,
    sep: str,
    auto_formula: str | None = None,
) -> str:
    if auto_formula is not None:
        auto_expression = _strip_formula_prefix(auto_formula)
        return (
            f'=IF(LEN(TRIM(({auto_expression})&""))>0{sep}'
            f'({auto_expression}){sep}'
            f'IF(OR({request_conditions}){sep}"依頼"{sep}""))'
        )
    return _build_request_formula(request_conditions)


def _build_request_header_formula(
    col_letter: str,
    sep: str,
    tool_row_min: int | None = None,
    tool_row_max: int | None = None,
    tool_row_step: int | None = None,
) -> str:
    if tool_row_min is None or tool_row_max is None or not tool_row_step:
        return ""
    conditions = (
        f"{_build_stepped_non_empty_count_formula(col_letter, tool_row_min, tool_row_max, tool_row_step, sep)}>0"
    )
    return _build_request_formula(conditions)


def _strip_formula_prefix(formula_text: str) -> str:
    if isinstance(formula_text, str) and formula_text.startswith("="):
        return formula_text[1:]
    return str(formula_text)


def _normalize_formula_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).replace("$", "").upper()


def _ensure_summary_formulas(ws):
    summary_col_start = column_index_from_string(SUMMARY_FORMULA_COL_START)
    summary_col_end = column_index_from_string(SUMMARY_FORMULA_COL_END)
    changed_refs = set()

    for col_idx in range(summary_col_start, summary_col_end + 1):
        col_letter = get_column_letter(col_idx)
        expected_row1_formula = _build_summary_formula(col_letter, 0)
        row1_cell = ws.cell(1, col_idx)
        if _normalize_formula_text(row1_cell.value) == _normalize_formula_text(
            expected_row1_formula
        ):
            continue

        for row_idx in range(1, 4):
            ws.cell(row_idx, col_idx).value = _build_summary_formula(
                col_letter,
                row_idx - 1,
            )
            changed_refs.add(f"{col_letter}{row_idx}")

    return changed_refs


def build_request_formulas(xlsx_path: str, out_path: str, cfg: dict, *, parent=None):
    sheet_name = cfg["sheet_name"]
    measure_no_col = cfg.get("measure_no_col", "A")
    measure_row_min = int(cfg.get("measure_row_min", 11))
    measure_row_step = int(cfg.get("measure_row_step", 3))
    tool_start_row = int(cfg.get("tool_start_row", 200))
    measure_row_max = int(cfg.get("measure_row_max", tool_start_row - 4))
    formula_arg_sep = str(cfg.get("formula_arg_sep", ",")).strip() or ","
    flag_col_start = column_index_from_string(REQUEST_OUTPUT_COL_START)
    flag_col_end = column_index_from_string(REQUEST_OUTPUT_COL_END)

    tool_name_col = cfg.get("tool_name_col", "E")
    tool_row_step = int(cfg.get("tool_row_step", measure_row_step))
    auto_data_start_row = int(cfg.get("auto_data_start_row", AUTO_DATA_START_ROW_DEFAULT))
    measure_no_to_data_index = _normalize_measure_to_index_map(
        cfg.get("measure_no_to_data_index", {})
    )

    tools = cfg["tools"]
    tool_to_measure_nos = cfg["tool_to_measure_nos"]

    wb = load_workbook(xlsx_path)
    wb_values = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"シート '{sheet_name}' が見つかりません。存在: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    ws_values = wb_values[sheet_name]
    changed_refs = set()

    no_col = column_index_from_string(measure_no_col)
    measure_no_to_row = {}
    measure_row_to_no = {}
    max_r = min(measure_row_max, ws.max_row or measure_row_max)
    for row_index in range(measure_row_min, max_r + 1):
        value = ws_values.cell(row_index, no_col).value
        if value is None:
            value = ws.cell(row_index, no_col).value
        if value is None:
            continue
        measure_no = _resolve_measure_no(value, row_index, measure_row_min, measure_row_step)
        if measure_no is None:
            continue
        measure_no_to_row[measure_no] = row_index
        measure_row_to_no[row_index] = measure_no

    tool_row = {}
    tool_name_col_index = column_index_from_string(tool_name_col)
    current_tool_row = tool_start_row
    for tool in tools:
        tool_cell = _get_writable_cell(ws, current_tool_row, tool_name_col_index)
        tool_cell.value = tool
        tool_row[tool] = tool_cell.row
        changed_refs.add(tool_cell.coordinate)
        current_tool_row += tool_row_step

    auto_data_label_cell = _get_writable_cell(ws, auto_data_start_row, tool_name_col_index)
    auto_data_label_cell.value = f"測定結果貼付は{auto_data_start_row}行から"
    changed_refs.add(auto_data_label_cell.coordinate)

    measure_row_to_tool_rows = {}
    missing_nos = []
    for tool, nos in tool_to_measure_nos.items():
        if tool not in tool_row:
            continue
        tool_row_index = tool_row[tool]
        for no in nos:
            measure_no = _try_extract_int(no)
            if measure_no is None:
                continue
            measure_row = measure_no_to_row.get(measure_no)
            if measure_row is None:
                missing_nos.append((tool, measure_no))
                continue
            measure_row_to_tool_rows.setdefault(measure_row, []).append(tool_row_index)

    all_tool_rows = sorted(tool_row.values())

    changed_refs.update(_ensure_summary_formulas(ws))

    written = 0
    target_found = 0
    for col_idx in range(flag_col_start, flag_col_end + 1):
        col_letter = get_column_letter(col_idx)
        header_cell = ws.cell(REQUEST_HEADER_ROW, col_idx)
        if _can_overwrite_with_formula(header_cell.value):
            header_formula = _build_request_header_formula(
                col_letter=col_letter,
                sep=formula_arg_sep,
                tool_row_min=all_tool_rows[0] if all_tool_rows else None,
                tool_row_max=all_tool_rows[-1] if all_tool_rows else None,
                tool_row_step=tool_row_step if all_tool_rows else None,
            )
            header_cell.value = header_formula
            changed_refs.add(header_cell.coordinate)

        for measure_row, tool_rows in measure_row_to_tool_rows.items():
            conditions = formula_arg_sep.join([f'{col_letter}${tool_row_index}<>""' for tool_row_index in tool_rows])
            measure_no = measure_row_to_no.get(measure_row)
            data_index = measure_no_to_data_index.get(measure_no)
            target_found += 1
            target_cell = ws.cell(measure_row, col_idx)
            if not _can_overwrite_with_formula(target_cell.value):
                continue
            if data_index is not None:
                auto_formula = _build_auto_data_formula(
                    col_letter=col_letter,
                    data_start_row=auto_data_start_row,
                    data_index=data_index,
                    sep=formula_arg_sep,
                )
                target_cell.value = _build_measure_row_formula(
                    conditions,
                    formula_arg_sep,
                    auto_formula,
                )
            else:
                target_cell.value = _build_measure_row_formula(conditions, formula_arg_sep)
            written += 1
            changed_refs.add(target_cell.coordinate)

        for measure_no, data_index in measure_no_to_data_index.items():
            measure_row = measure_no_to_row.get(measure_no)
            if measure_row is None or measure_row in measure_row_to_tool_rows:
                continue
            target_found += 1
            target_cell = ws.cell(measure_row, col_idx)
            if not _can_overwrite_with_formula(target_cell.value):
                continue
            auto_formula = _build_auto_data_formula(
                col_letter=col_letter,
                data_start_row=auto_data_start_row,
                data_index=data_index,
                sep=formula_arg_sep,
            )
            target_cell.value = f"={auto_formula}"
            written += 1
            changed_refs.add(target_cell.coordinate)

    if target_found == 0:
        raise ValueError(
            "依頼/自動測定データの書き込み対象が1件も見つかりませんでした。\n\n"
            f"- 読み取れた測定No件数: {len(measure_no_to_row)}\n"
            f"- 工具件数: {len(tools)}\n"
            f"- 逆引き対象の測定行件数: {len(measure_row_to_tool_rows)}\n"
            "- 出力列: L～SR（固定）\n"
            f"- 指定Noが見つからない例(先頭10件): {missing_nos[:10]}\n\n"
            "原因候補:\n"
            "1) 測定No列/行範囲が実シートと違う\n"
            "2) 測定Noが数式で、キャッシュ値が未保存（Excelで一度保存してから再実行）\n"
            "3) tool_to_measure_nos のNoがシートに存在しない"
        )

    normalized_refs = _normalize_single_cell_array_formulas_in_column(
        ws,
        "B",
        row_start=measure_row_min,
        row_end=measure_row_max,
    )
    if normalized_refs:
        changed_refs.update(normalized_refs)
        print(f"[info] B列の単一セル配列数式を通常数式へ変換: {len(normalized_refs)}件")

    saved_path = _finalize_modified_workbook(
        wb,
        wb_values,
        source_xlsx_path=xlsx_path,
        out_path=out_path,
        sheet_name=sheet_name,
        changed_refs=changed_refs,
        parent=parent,
    )
    _force_excel_recalc_and_save(saved_path)
    return saved_path


def _parse_int_list(text: str):
    if not text or not text.strip():
        return []
    parts = re.split(r"[,\s、，;；]+", text.strip())
    nums = []
    for part in parts:
        if not part:
            continue
        value = _try_extract_int(part)
        if value is None:
            raise ValueError(f"測定Noに整数以外の入力が含まれています: '{part}'")
        nums.append(value)
    return nums


def write_measurement_not_required(
    xlsx_path: str,
    out_path: str,
    cfg: dict,
    target_nos: list | None = None,
    *,
    parent=None,
):
    if target_nos is None:
        target_nos = []

    sheet_name = cfg.get("sheet_name", "工程内検査シート")
    measure_no_col = cfg.get("measure_no_col", "A")
    tool_start_row = int(cfg.get("tool_start_row", 200))
    not_required_row = int(cfg.get("not_required_row", tool_start_row - 3))
    measure_row_min = int(cfg.get("measure_row_min", 11))
    measure_row_step = int(cfg.get("measure_row_step", 3))
    measure_row_max = int(cfg.get("measure_row_max", tool_start_row - 4))
    formula_arg_sep = str(cfg.get("formula_arg_sep", ",")).strip() or ","

    wb = load_workbook(xlsx_path)
    wb_values = load_workbook(xlsx_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"シート '{sheet_name}' が見つかりません。存在: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    ws_values = wb_values[sheet_name]
    changed_refs = set()

    no_col = column_index_from_string(measure_no_col)
    measure_no_to_row = {}
    max_r = min(measure_row_max, ws.max_row or measure_row_max)
    debug_info = []

    for row_index in range(measure_row_min, max_r + 1):
        value = ws_values.cell(row_index, no_col).value
        cell_obj = ws.cell(row_index, no_col)
        formula = None

        if value is None:
            value = cell_obj.value
            if isinstance(value, str) and value.startswith("="):
                formula = value

        if value is None:
            continue

        measure_no = _resolve_measure_no(value, row_index, measure_row_min, measure_row_step)

        if len(debug_info) < 10:
            debug_info.append(
                f"行{row_index}: 値={value!r}, 型={type(value).__name__}, 数式={formula}, 解決No={measure_no}"
            )

        if measure_no is None:
            continue
        measure_no_to_row[measure_no] = row_index

    target_row = not_required_row

    e_col = column_index_from_string("E")
    target_e_cell = _get_writable_cell(ws, target_row, e_col)
    target_e_cell.value = "測定不要"
    changed_refs.add(target_e_cell.coordinate)

    flag_col_start = column_index_from_string("L")
    flag_col_end = column_index_from_string("SR")

    written_count = 0
    for no in target_nos:
        row_index = measure_no_to_row.get(no)
        if row_index is None:
            continue

        for col_idx in range(flag_col_start, flag_col_end + 1):
            col_letter = get_column_letter(col_idx)
            target_cell = f"{col_letter}{target_row}"
            current_cell = ws.cell(row_index, col_idx)
            if not _can_overwrite_with_formula(current_cell.value):
                continue
            current_cell.value = _build_not_required_overlay_formula(
                current_cell.value,
                target_cell,
                formula_arg_sep,
            )
            changed_refs.add(current_cell.coordinate)
        written_count += 1

    if written_count == 0 and target_nos:
        available_nos = sorted(measure_no_to_row.keys())
        available_nos_str = ", ".join(map(str, available_nos[:20]))
        if len(available_nos) > 20:
            available_nos_str += f" ... (他{len(available_nos) - 20}件)"

        debug_str = "\n".join(debug_info)

        raise ValueError(
            f"指定したNo.の行が見つかりませんでした。\n\n"
            f"【指定したNo.】\n{target_nos}\n\n"
            f"【読み取れた測定No】\n{available_nos_str}\n\n"
            f"【設定】\n"
            f"- 測定No列: {measure_no_col}列\n"
            f"- 測定行範囲: {measure_row_min}〜{measure_row_max}行目\n"
            f"- 読み取れた測定No件数: {len(measure_no_to_row)}件\n\n"
            f"【デバッグ情報（先頭10行）】\n{debug_str}\n\n"
            f"※測定Noが数式の場合、Excelで一度ファイルを開いて保存してから再実行してください。"
        )

    normalized_refs = _normalize_single_cell_array_formulas_in_column(
        ws,
        "B",
        row_start=measure_row_min,
        row_end=measure_row_max,
    )
    if normalized_refs:
        changed_refs.update(normalized_refs)
        print(f"[info] B列の単一セル配列数式を通常数式へ変換: {len(normalized_refs)}件")

    saved_path = _finalize_modified_workbook(
        wb,
        wb_values,
        source_xlsx_path=xlsx_path,
        out_path=out_path,
        sheet_name=sheet_name,
        changed_refs=changed_refs,
        parent=parent,
    )
    _force_excel_recalc_and_save(saved_path)
    return saved_path
